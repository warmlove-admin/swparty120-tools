import argparse
import csv
import json
import os
import re
import shutil
import sys
import time
from dataclasses import asdict, dataclass
from datetime import datetime, timedelta, time as dtime
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
CODE_RE = re.compile(r"^[1-9][0-9]{3}$")

HEADER_MAP = {
    "watch_level": "觀察等級",
    "mode_count": "上榜模式數",
    "code": "股票代號",
    "name": "股票名稱",
    "exchange": "市場",
    "modes": "上榜模式",
    "levels": "各模式等級",
    "avg_score": "平均分數",
    "min_score": "最低分數",
    "entry_price": "參考進場價",
    "nearest_target_price": "最近停利價",
    "nearest_stop_price": "最近停損價",
    "min_net_profit": "最低預估淨利",
    "min_rr_ratio": "最低風報比",
    "return_pct": "漲跌幅%",
    "volume": "成交量",
    "reasons": "入選原因",
    "mode": "模式",
    "level": "等級",
    "score": "分數",
    "target_price": "停利價",
    "stop_price": "停損價",
    "lot_amount": "單張金額",
    "net_profit_at_target": "停利預估淨利",
    "loss_at_stop": "停損預估虧損",
    "rr_ratio": "風報比",
    "gap_pct": "開盤跳空%",
    "fade_pct": "開盤後漲跌%",
    "spread_pct": "買賣價差%",
    "total_amount": "成交金額",
    "liquidity_score": "流動性分數",
    "momentum_score": "動能分數",
    "technical_score": "技術分數",
    "risk_score": "風險分數",
    "exclude_reasons": "排除原因",
    "ts": "資料時間",
    "day_trade": "可否當沖",
    "category": "產業代碼",
    "reference": "參考價",
    "limit_up": "漲停價",
    "limit_down": "跌停價",
    "open": "開盤價",
    "high": "最高價",
    "low": "最低價",
    "close": "成交/收盤價",
    "buy_price": "買一價",
    "sell_price": "賣一價",
}

MODE_MAP = {"daytrade": "當沖", "overnight": "隔日沖", "weekly": "週沖", "pre_breakout": "起漲點"}
WATCH_MAP = {
    "core_watch_strong": "核心觀察-強",
    "core_watch_with_breakout": "核心觀察-含起漲",
    "core_watch": "核心觀察",
    "priority_watch": "優先觀察",
}


@dataclass
class Settings:
    max_price: float = 200.0
    max_lot_amount: float = 200_000.0
    min_volume: int = 1000
    max_spread_pct: float = 0.6
    commission_rate: float = 0.001425
    commission_discount: float = 0.6
    min_commission: float = 20.0
    intraday_tax_rate: float = 0.0015
    regular_tax_rate: float = 0.003
    daytrade_target_pct: float = 2.5
    daytrade_stop_pct: float = 0.6
    daytrade_min_net_profit: float = 800.0
    overnight_target_pct: float = 5.0
    overnight_stop_pct: float = 1.5
    overnight_min_net_profit: float = 1500.0
    weekly_target_pct: float = 8.0
    weekly_stop_pct: float = 3.0
    weekly_min_net_profit: float = 2500.0
    pre_breakout_target_pct: float = 6.5
    pre_breakout_stop_pct: float = 2.0
    pre_breakout_min_net_profit: float = 1200.0
    pre_breakout_min_gap_pct: float = 0.5
    pre_breakout_max_gap_pct: float = 5.5
    min_rr_ratio: float = 2.0
    history_pool_size: int = 80
    kbar_days: int = 70


@dataclass
class Candidate:
    mode: str
    code: str
    name: str
    exchange: str
    level: str
    score: float
    entry_price: float
    target_price: float
    stop_price: float
    lot_amount: float
    net_profit_at_target: float
    loss_at_stop: float
    rr_ratio: float
    return_pct: float
    gap_pct: float
    fade_pct: float
    spread_pct: float
    volume: int
    total_amount: float
    liquidity_score: int
    momentum_score: int
    technical_score: int
    risk_score: int
    reasons: str


def env_float(name, default):
    value = os.getenv(name)
    if value in (None, ""):
        return default
    return float(value)


def env_int(name, default):
    value = os.getenv(name)
    if value in (None, ""):
        return default
    return int(value)


def load_settings():
    return Settings(
        max_price=env_float("STOCK_MAX_PRICE", Settings.max_price),
        max_lot_amount=env_float("STOCK_MAX_LOT_AMOUNT", Settings.max_lot_amount),
        min_volume=env_int("STOCK_MIN_VOLUME", Settings.min_volume),
        max_spread_pct=env_float("STOCK_MAX_SPREAD_PCT", Settings.max_spread_pct),
        commission_rate=env_float("STOCK_COMMISSION_RATE", Settings.commission_rate),
        commission_discount=env_float("STOCK_COMMISSION_DISCOUNT", Settings.commission_discount),
        min_commission=env_float("STOCK_MIN_COMMISSION", Settings.min_commission),
        intraday_tax_rate=env_float("STOCK_INTRADAY_TAX_RATE", Settings.intraday_tax_rate),
        regular_tax_rate=env_float("STOCK_REGULAR_TAX_RATE", Settings.regular_tax_rate),
        daytrade_target_pct=env_float("DAYTRADE_TARGET_PCT", Settings.daytrade_target_pct),
        daytrade_stop_pct=env_float("DAYTRADE_STOP_PCT", Settings.daytrade_stop_pct),
        daytrade_min_net_profit=env_float("DAYTRADE_MIN_NET_PROFIT", Settings.daytrade_min_net_profit),
        overnight_target_pct=env_float("OVERNIGHT_TARGET_PCT", Settings.overnight_target_pct),
        overnight_stop_pct=env_float("OVERNIGHT_STOP_PCT", Settings.overnight_stop_pct),
        overnight_min_net_profit=env_float("OVERNIGHT_MIN_NET_PROFIT", Settings.overnight_min_net_profit),
        weekly_target_pct=env_float("WEEKLY_TARGET_PCT", Settings.weekly_target_pct),
        weekly_stop_pct=env_float("WEEKLY_STOP_PCT", Settings.weekly_stop_pct),
        weekly_min_net_profit=env_float("WEEKLY_MIN_NET_PROFIT", Settings.weekly_min_net_profit),
        pre_breakout_target_pct=env_float("PRE_BREAKOUT_TARGET_PCT", Settings.pre_breakout_target_pct),
        pre_breakout_stop_pct=env_float("PRE_BREAKOUT_STOP_PCT", Settings.pre_breakout_stop_pct),
        pre_breakout_min_net_profit=env_float("PRE_BREAKOUT_MIN_NET_PROFIT", Settings.pre_breakout_min_net_profit),
        pre_breakout_min_gap_pct=env_float("PRE_BREAKOUT_MIN_GAP_PCT", Settings.pre_breakout_min_gap_pct),
        pre_breakout_max_gap_pct=env_float("PRE_BREAKOUT_MAX_GAP_PCT", Settings.pre_breakout_max_gap_pct),
        min_rr_ratio=env_float("STOCK_MIN_RR_RATIO", Settings.min_rr_ratio),
        history_pool_size=env_int("STOCK_HISTORY_POOL_SIZE", Settings.history_pool_size),
        kbar_days=env_int("STOCK_KBAR_DAYS", Settings.kbar_days),
    )


def env_required(name):
    value = os.getenv(name)
    if not value:
        raise RuntimeError(f"Missing required environment variable: {name}")
    return value


def to_float(value, default=0.0):
    try:
        if value is None:
            return default
        return float(value)
    except Exception:
        return default


def to_exchange(value):
    text = str(value)
    if "." in text:
        text = text.rsplit(".", 1)[-1]
    return text.replace("Exchange.", "")


def load_api():
    load_dotenv(BASE_DIR / ".env", encoding="utf-8-sig")
    import shioaji as sj

    simulation = os.getenv("SHIOAJI_SIMULATION", "true").lower() != "false"
    api = sj.Shioaji(simulation=simulation)
    api.login(api_key=env_required("SHIOAJI_API_KEY"), secret_key=env_required("SHIOAJI_SECRET_KEY"))
    return api, simulation


def safe_logout(api):
    try:
        api.logout()
    except Exception as exc:
        print(f"logout skipped: {exc}", file=sys.stderr)


def iter_stock_contracts(api):
    for group_name in ("TSE", "OTC"):
        group = getattr(api.Contracts.Stocks, group_name)
        for contract in group:
            code = getattr(contract, "code", "")
            if CODE_RE.match(code):
                yield contract


def chunks(items, size):
    for index in range(0, len(items), size):
        yield items[index : index + size]


def snapshot_to_dict(snapshot, contract):
    return {
        "code": getattr(contract, "code", ""),
        "name": getattr(contract, "name", ""),
        "exchange": to_exchange(getattr(contract, "exchange", "")),
        "day_trade": str(getattr(contract, "day_trade", "")),
        "category": getattr(contract, "category", ""),
        "reference": to_float(getattr(contract, "reference", 0)),
        "limit_up": to_float(getattr(contract, "limit_up", 0)),
        "limit_down": to_float(getattr(contract, "limit_down", 0)),
        "open": to_float(getattr(snapshot, "open", 0)),
        "high": to_float(getattr(snapshot, "high", 0)),
        "low": to_float(getattr(snapshot, "low", 0)),
        "close": to_float(getattr(snapshot, "close", 0)),
        "buy_price": to_float(getattr(snapshot, "buy_price", 0)),
        "sell_price": to_float(getattr(snapshot, "sell_price", 0)),
        "volume": int(to_float(getattr(snapshot, "total_volume", 0))),
        "total_amount": to_float(getattr(snapshot, "total_amount", 0)),
        "ts": datetime.now().isoformat(timespec="seconds"),
    }


def fetch_snapshots(api):
    contracts = list(iter_stock_contracts(api))
    by_code = {contract.code: contract for contract in contracts}
    rows = []
    for batch in chunks(contracts, 200):
        try:
            snapshots = api.snapshots(batch)
        except Exception as exc:
            print(f"snapshot batch failed: {exc}", file=sys.stderr)
            continue
        for snapshot in snapshots:
            contract = by_code.get(str(getattr(snapshot, "code", "")))
            if contract:
                rows.append(snapshot_to_dict(snapshot, contract))
        time.sleep(0.1)
    return rows, by_code


def fee(amount, settings):
    return max(settings.min_commission, amount * settings.commission_rate * settings.commission_discount)


def trade_math(entry, target, stop, mode, settings):
    shares = 1000
    tax_rate = settings.intraday_tax_rate if mode == "daytrade" else settings.regular_tax_rate
    buy_amount = entry * shares
    target_amount = target * shares
    stop_amount = stop * shares

    buy_fee = fee(buy_amount, settings)
    sell_fee_target = fee(target_amount, settings)
    sell_fee_stop = fee(stop_amount, settings)
    tax_target = target_amount * tax_rate
    tax_stop = stop_amount * tax_rate

    net_profit = target_amount - buy_amount - buy_fee - sell_fee_target - tax_target
    stop_loss = buy_amount - stop_amount + buy_fee + sell_fee_stop + tax_stop
    rr = net_profit / stop_loss if stop_loss > 0 else 0
    return round(net_profit, 0), round(stop_loss, 0), round(rr, 2)


def common_exclusions(row, settings):
    reasons = []
    close = row["close"]
    reference = row["reference"]
    buy_price = row["buy_price"]
    sell_price = row["sell_price"]
    spread_pct = (sell_price - buy_price) / close * 100 if close and buy_price and sell_price else 999

    if close <= 0 or reference <= 0:
        reasons.append("invalid_quote")
    if row["volume"] < settings.min_volume:
        reasons.append("volume_too_low")
    if spread_pct > settings.max_spread_pct:
        reasons.append("spread_too_wide")
    if close > settings.max_price:
        reasons.append("price_above_limit")
    if close * 1000 > settings.max_lot_amount:
        reasons.append("one_lot_amount_above_limit")
    return reasons


def mode_trade_plan(row, mode, settings):
    entry = row["close"]
    if mode == "daytrade":
        target_pct = settings.daytrade_target_pct
        stop_pct = settings.daytrade_stop_pct
        min_net = settings.daytrade_min_net_profit
    elif mode == "overnight":
        target_pct = settings.overnight_target_pct
        stop_pct = settings.overnight_stop_pct
        min_net = settings.overnight_min_net_profit
    elif mode == "weekly":
        target_pct = settings.weekly_target_pct
        stop_pct = settings.weekly_stop_pct
        min_net = settings.weekly_min_net_profit
    else:
        target_pct = settings.pre_breakout_target_pct
        stop_pct = settings.pre_breakout_stop_pct
        min_net = settings.pre_breakout_min_net_profit

    target = round(entry * (1 + target_pct / 100), 2)
    stop = round(entry * (1 - stop_pct / 100), 2)
    net_profit, stop_loss, rr = trade_math(entry, target, stop, mode, settings)
    reasons = []
    if net_profit < min_net:
        reasons.append("net_profit_below_min")
    if rr < settings.min_rr_ratio:
        reasons.append("rr_below_min")
    return target, stop, net_profit, stop_loss, rr, reasons


def snapshot_base_scores(row):
    close = row["close"]
    open_price = row["open"]
    high = row["high"]
    low = row["low"]
    reference = row["reference"]
    buy_price = row["buy_price"]
    sell_price = row["sell_price"]
    volume = row["volume"]

    return_pct = (close - reference) / reference * 100 if reference else 0
    gap_pct = (open_price - reference) / reference * 100 if reference else 0
    fade_pct = (close - open_price) / open_price * 100 if open_price else 0
    spread_pct = (sell_price - buy_price) / close * 100 if close and buy_price and sell_price else 999
    range_pct = (high - low) / reference * 100 if reference else 0

    liquidity = 0
    liquidity += 8 if volume >= 10000 else 6 if volume >= 5000 else 4 if volume >= 2000 else 2
    liquidity += 6 if row["total_amount"] >= 500_000_000 else 4 if row["total_amount"] >= 150_000_000 else 2
    liquidity += 4 if spread_pct <= 0.2 else 3 if spread_pct <= 0.4 else 1

    momentum = 0
    momentum += 5 if gap_pct >= 2 else 3 if gap_pct >= 1 else 1 if gap_pct > 0 else 0
    momentum += 5 if return_pct >= 4 else 4 if return_pct >= 3 else 3 if return_pct >= 2 else 1 if return_pct > 0 else 0
    if gap_pct > 0 and fade_pct >= 0 and return_pct > 1:
        momentum += 3

    technical = 0
    if close > reference:
        technical += 3
    if close > open_price:
        technical += 2
    if 2 <= range_pct <= 6:
        technical += 2
    if fade_pct >= 0:
        technical += 2
    if spread_pct <= 0.2:
        technical += 1

    risk = 5
    if range_pct > 6:
        risk -= 2
    if spread_pct > 0.4:
        risk -= 2

    reasons = []
    if close > open_price:
        reasons.append("above_open")
    if high > low and (high - close) / (high - low) <= 0.25:
        reasons.append("near_day_high")
    if return_pct >= 4 and volume >= 3000:
        reasons.append("strong_momentum")

    return {
        "return_pct": return_pct,
        "gap_pct": gap_pct,
        "fade_pct": fade_pct,
        "spread_pct": spread_pct,
        "liquidity": liquidity,
        "momentum": min(15, momentum),
        "technical": min(10, technical),
        "risk": max(0, risk),
        "reasons": reasons,
    }


def make_candidate(row, mode, settings, history=None):
    if mode == "daytrade" and "Yes" not in row["day_trade"]:
        return None, ["not_day_trade"]

    reasons = common_exclusions(row, settings)
    target, stop, net_profit, stop_loss, rr, trade_reasons = mode_trade_plan(row, mode, settings)
    reasons.extend(trade_reasons)
    if reasons:
        return None, reasons

    scores = snapshot_base_scores(row)
    liquidity = scores["liquidity"]
    momentum = scores["momentum"]
    technical = scores["technical"]
    risk = scores["risk"]
    reason_tags = list(scores["reasons"])

    if mode == "overnight" and history:
        add, tags = overnight_history_score(history)
        technical += add
        reason_tags.extend(tags)
    elif mode == "weekly" and history:
        add, tags = weekly_history_score(history)
        technical += add
        reason_tags.extend(tags)
    elif mode == "pre_breakout":
        if not history:
            return None, ["not_enough_history"]
        add, tags, blockers = pre_breakout_history_score(row, history, settings)
        if blockers:
            return None, blockers
        technical += add
        reason_tags.extend(tags)

    max_raw = 73 if mode == "daytrade" else 88 if mode == "pre_breakout" else 80
    raw = liquidity + momentum + min(20, technical) + risk
    if mode == "daytrade":
        raw += 15 if row["close"] > row["open"] else 5
    elif mode == "pre_breakout":
        raw += 12 if row["close"] >= row["open"] else 0
    else:
        raw += 10 if scores["return_pct"] > 0 else 0

    score = round(raw / max_raw * 100, 1)
    level = "A" if score >= 80 and liquidity >= 14 and rr >= settings.min_rr_ratio else "B" if score >= 70 else "C"
    if level == "C":
        return None, ["score_below_b"]

    return Candidate(
        mode=mode,
        code=row["code"],
        name=row["name"],
        exchange=row["exchange"],
        level=level,
        score=score,
        entry_price=round(row["close"], 4),
        target_price=target,
        stop_price=stop,
        lot_amount=round(row["close"] * 1000, 0),
        net_profit_at_target=net_profit,
        loss_at_stop=stop_loss,
        rr_ratio=rr,
        return_pct=round(scores["return_pct"], 2),
        gap_pct=round(scores["gap_pct"], 2),
        fade_pct=round(scores["fade_pct"], 2),
        spread_pct=round(scores["spread_pct"], 3),
        volume=row["volume"],
        total_amount=round(row["total_amount"], 0),
        liquidity_score=liquidity,
        momentum_score=momentum,
        technical_score=min(20, technical),
        risk_score=risk,
        reasons=",".join(reason_tags[:8]),
    ), []


def daily_bars_from_kbars(kbars):
    grouped = {}
    for ts, open_, high, low, close, volume in zip(
        kbars.ts, kbars.Open, kbars.High, kbars.Low, kbars.Close, kbars.Volume
    ):
        day = datetime.fromtimestamp(ts / 1_000_000_000).date().isoformat()
        bucket = grouped.setdefault(
            day,
            {"date": day, "open": open_, "high": high, "low": low, "close": close, "volume": 0},
        )
        bucket["high"] = max(bucket["high"], high)
        bucket["low"] = min(bucket["low"], low)
        bucket["close"] = close
        bucket["volume"] += volume
    return [grouped[key] for key in sorted(grouped)]


def avg(values):
    return sum(values) / len(values) if values else 0


def fetch_history_features(api, contracts, days):
    end = datetime.now().date()
    start = end - timedelta(days=days)
    features = {}
    for contract in contracts:
        try:
            kbars = api.kbars(contract, start=start.isoformat(), end=end.isoformat())
            bars = daily_bars_from_kbars(kbars)
            if len(bars) >= 10:
                features[contract.code] = bars
        except Exception as exc:
            print(f"kbar failed {contract.code}: {exc}", file=sys.stderr)
            text = str(exc)
            if "SessionNotEstablished" in text or "NotReady" in text:
                break
        time.sleep(0.03)
    return features


def overnight_history_score(bars):
    tags = []
    closes = [bar["close"] for bar in bars]
    vols = [bar["volume"] for bar in bars]
    last = bars[-1]
    ma5 = avg(closes[-5:])
    ma10 = avg(closes[-10:])
    vol5 = avg(vols[-5:])
    score = 0
    if last["close"] > ma5:
        score += 4
        tags.append("above_ma5")
    if last["close"] > ma10:
        score += 3
        tags.append("above_ma10")
    if ma5 > ma10:
        score += 3
        tags.append("ma5_gt_ma10")
    if vol5 and last["volume"] > vol5 * 1.3:
        score += 4
        tags.append("volume_expansion")
    if last["high"] > last["low"] and (last["high"] - last["close"]) / (last["high"] - last["low"]) <= 0.25:
        score += 3
        tags.append("daily_close_near_high")
    return score, tags


def weekly_history_score(bars):
    tags = []
    closes = [bar["close"] for bar in bars]
    vols = [bar["volume"] for bar in bars]
    last = bars[-1]
    ma5 = avg(closes[-5:])
    ma10 = avg(closes[-10:])
    ma20 = avg(closes[-20:]) if len(closes) >= 20 else ma10
    vol20 = avg(vols[-20:]) if len(vols) >= 20 else avg(vols)
    score = 0
    if last["close"] > ma5 > ma10:
        score += 5
        tags.append("short_trend_up")
    if ma5 > ma10 > ma20:
        score += 5
        tags.append("ma_alignment")
    if closes[-1] > closes[-5]:
        score += 3
        tags.append("five_day_price_up")
    if vol20 and avg(vols[-5:]) > vol20 * 1.2:
        score += 4
        tags.append("five_day_volume_up")
    if last["close"] > ma20:
        score += 3
        tags.append("above_ma20")
    return score, tags


def calc_kd(bars, period=9):
    k_value = 50.0
    d_value = 50.0
    for index in range(len(bars)):
        start = max(0, index - period + 1)
        window = bars[start : index + 1]
        high = max(bar["high"] for bar in window)
        low = min(bar["low"] for bar in window)
        close = bars[index]["close"]
        rsv = 50.0 if high == low else (close - low) / (high - low) * 100
        k_value = k_value * 2 / 3 + rsv / 3
        d_value = d_value * 2 / 3 + k_value / 3
    return k_value, d_value


def pre_breakout_history_score(row, bars, settings):
    if len(bars) < 20:
        return 0, [], ["not_enough_history"]

    # Shioaji kbars usually includes today's incomplete bar during trading hours.
    history = bars[:-1] if bars[-1]["date"] == datetime.now().date().isoformat() and len(bars) > 20 else bars
    if len(history) < 20:
        return 0, [], ["not_enough_history"]

    open_price = row["open"]
    close = row["close"]
    reference = row["reference"]
    volume = row["volume"]
    gap_pct = (open_price - reference) / reference * 100 if reference else 0

    blockers = []
    if gap_pct < settings.pre_breakout_min_gap_pct:
        blockers.append("gap_too_small")
    if gap_pct > settings.pre_breakout_max_gap_pct:
        blockers.append("gap_too_large")

    closes = [bar["close"] for bar in history]
    vols = [bar["volume"] for bar in history]
    prev = history[-1]
    ma5 = avg(closes[-5:])
    ma10 = avg(closes[-10:])
    ma20 = avg(closes[-20:])
    ma60 = avg(closes[-60:]) if len(closes) >= 60 else avg(closes)
    vol5 = avg(vols[-5:])

    score = 0
    tags = []

    if open_price > ma5:
        score += 5
        tags.append("open_above_ma5")
    if open_price > ma10:
        score += 5
        tags.append("open_above_ma10")
    if not (open_price > ma5 and open_price > ma10):
        blockers.append("not_above_short_ma")

    ma20_distance = (ma20 - open_price) / open_price * 100 if open_price else 999
    ma60_distance = (ma60 - open_price) / open_price * 100 if open_price else 999
    if open_price >= ma20:
        score += 5
        tags.append("open_above_ma20")
    elif 0 <= ma20_distance <= 3:
        score += 4
        tags.append("near_ma20_breakout")
    if open_price >= ma60:
        score += 5
        tags.append("open_above_ma60")
    elif 0 <= ma60_distance <= 5:
        score += 4
        tags.append("near_ma60_breakout")

    prev_high_distance = (prev["high"] - open_price) / open_price * 100 if open_price else 999
    if open_price > prev["high"]:
        score += 6
        tags.append("open_breaks_prev_high")
    elif 0 <= prev_high_distance <= 3:
        score += 5
        tags.append("near_prev_high")
    elif prev_high_distance > 5:
        blockers.append("too_far_from_prev_high")

    if vol5 and volume > vol5 * 1.5:
        score += 5
        tags.append("volume_expansion")
    elif vol5 and volume > vol5:
        score += 2
        tags.append("volume_above_avg")

    kd_bars = history[-8:] + [
        {
            "date": datetime.now().date().isoformat(),
            "open": row["open"],
            "high": max(row["high"], row["close"]),
            "low": min(row["low"], row["close"]),
            "close": row["close"],
            "volume": row["volume"],
        }
    ]
    k_value, d_value = calc_kd(kd_bars)
    if k_value > 60 and k_value > d_value and d_value > 50:
        score += 8
        tags.append("kd_confirmed")
    elif 30 <= k_value <= 60 and k_value > d_value:
        score += 3
        tags.append("kd_early_turn")
    elif k_value < d_value and close < ma5:
        blockers.append("kd_weak_and_below_ma5")
    if k_value > 90 and gap_pct > 4:
        tags.append("kd_overheat")

    if close >= open_price:
        score += 4
        tags.append("holds_open")
    else:
        blockers.append("falls_below_open")

    return score, tags, blockers


def write_csv(path, rows, fieldnames):
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def localized_value(header, value):
    if value is None:
        return value
    if header in ("modes", "mode"):
        return ",".join(MODE_MAP.get(part.strip(), part.strip()) for part in str(value).split(","))
    if header == "watch_level":
        return WATCH_MAP.get(str(value), value)
    return value


def autosize_sheet(ws, max_width=42):
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        width = 10
        for cell in column[:300]:
            if cell.value is not None:
                width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(bottom=thin)
    ws.freeze_panes = "A2"
    if ws.max_row and ws.max_column:
        ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top")
    autosize_sheet(ws)


def write_xlsx_sheet(wb, title, rows, fields):
    ws = wb.create_sheet(title[:31])
    ws.append([HEADER_MAP.get(field, field) for field in fields])
    for row in rows:
        ws.append([localized_value(field, row.get(field)) for field in fields])
    style_sheet(ws)


def write_excel_report(output_dir, snapshot_rows, groups, overlap_rows, excluded, output_name="stock_screen_report.xlsx"):
    wb = Workbook()
    wb.remove(wb.active)

    summary = wb.create_sheet("說明")
    summary.append(["項目", "內容"])
    summary.append(["來源資料夾", str(output_dir)])
    summary.append(["主要分頁", "先看「重疊名單」，再看當沖/隔日沖/週沖"])
    summary.append(["注意", "重疊名單是觀察優先級，不是直接買進訊號"])
    style_sheet(summary)

    overlap_fields = [
        "watch_level",
        "mode_count",
        "code",
        "name",
        "exchange",
        "modes",
        "levels",
        "avg_score",
        "min_score",
        "entry_price",
        "nearest_target_price",
        "nearest_stop_price",
        "min_net_profit",
        "min_rr_ratio",
        "return_pct",
        "volume",
        "reasons",
    ]
    candidate_fields = list(Candidate.__annotations__.keys())
    snapshot_fields = [
        "ts",
        "code",
        "name",
        "exchange",
        "day_trade",
        "category",
        "reference",
        "limit_up",
        "limit_down",
        "open",
        "high",
        "low",
        "close",
        "buy_price",
        "sell_price",
        "volume",
        "total_amount",
    ]
    excluded_fields = list(excluded[0].keys()) if excluded else ["mode", "code", "name", "exclude_reasons"]

    write_xlsx_sheet(wb, "重疊名單", overlap_rows, overlap_fields)
    write_xlsx_sheet(wb, "起漲點", [asdict(row) for row in groups["pre_breakout"]], candidate_fields)
    write_xlsx_sheet(wb, "當沖", [asdict(row) for row in groups["daytrade"]], candidate_fields)
    write_xlsx_sheet(wb, "隔日沖", [asdict(row) for row in groups["overnight"]], candidate_fields)
    write_xlsx_sheet(wb, "週沖", [asdict(row) for row in groups["weekly"]], candidate_fields)
    write_xlsx_sheet(wb, "排除清單", excluded, excluded_fields)
    write_xlsx_sheet(wb, "原始快照", snapshot_rows, snapshot_fields)
    target = output_dir / output_name
    try:
        wb.save(target)
    except PermissionError:
        fallback = output_dir / f"{target.stem}_{datetime.now():%H%M%S}.xlsx"
        wb.save(fallback)
        print(f"xlsx target locked; saved fallback={fallback}", file=sys.stderr)


def copy_labeled_outputs(output_dir, label):
    if not label:
        return
    names = [
        "pre_breakout_candidates.csv",
        "pre_breakout_candidates.json",
        "overlap_candidates.csv",
        "overlap_candidates.json",
        "daytrade_candidates.csv",
        "overnight_candidates.csv",
        "weekly_candidates.csv",
        "screening_report.md",
    ]
    for name in names:
        source = output_dir / name
        if not source.exists():
            continue
        target = source.with_name(f"{source.stem}_{label}{source.suffix}")
        shutil.copy2(source, target)


def write_candidates(output_dir, filename, rows):
    dicts = [asdict(row) for row in rows]
    fields = list(Candidate.__annotations__.keys())
    write_csv(output_dir / f"{filename}.csv", dicts, fields)
    (output_dir / f"{filename}.json").write_text(
        json.dumps(dicts, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def build_overlap_candidates(groups):
    by_code = {}
    for mode, rows in groups.items():
        for row in rows:
            item = by_code.setdefault(
                row.code,
                {
                    "code": row.code,
                    "name": row.name,
                    "exchange": row.exchange,
                    "modes": [],
                    "levels": [],
                    "scores": [],
                    "entry_prices": [],
                    "target_prices": [],
                    "stop_prices": [],
                    "net_profits": [],
                    "rr_ratios": [],
                    "return_pct": row.return_pct,
                    "volume": row.volume,
                    "reasons": set(),
                },
            )
            item["modes"].append(mode)
            item["levels"].append(row.level)
            item["scores"].append(row.score)
            item["entry_prices"].append(row.entry_price)
            item["target_prices"].append(row.target_price)
            item["stop_prices"].append(row.stop_price)
            item["net_profits"].append(row.net_profit_at_target)
            item["rr_ratios"].append(row.rr_ratio)
            item["reasons"].update(tag for tag in row.reasons.split(",") if tag)

    overlap = []
    for item in by_code.values():
        mode_count = len(item["modes"])
        if mode_count < 2:
            continue
        watch_level = "core_watch" if mode_count >= 3 else "priority_watch"
        if mode_count >= 4:
            watch_level = "core_watch_with_breakout"
        if mode_count >= 3 and all(level == "A" for level in item["levels"]):
            watch_level = "core_watch_strong"
        overlap.append(
            {
                "watch_level": watch_level,
                "mode_count": mode_count,
                "code": item["code"],
                "name": item["name"],
                "exchange": item["exchange"],
                "modes": ",".join(item["modes"]),
                "levels": ",".join(item["levels"]),
                "avg_score": round(avg(item["scores"]), 1),
                "min_score": round(min(item["scores"]), 1),
                "entry_price": round(avg(item["entry_prices"]), 2),
                "nearest_target_price": round(min(item["target_prices"]), 2),
                "nearest_stop_price": round(max(item["stop_prices"]), 2),
                "min_net_profit": round(min(item["net_profits"]), 0),
                "min_rr_ratio": round(min(item["rr_ratios"]), 2),
                "return_pct": item["return_pct"],
                "volume": item["volume"],
                "reasons": ",".join(sorted(item["reasons"])[:10]),
            }
        )
    overlap.sort(key=lambda row: (row["mode_count"], row["avg_score"], row["min_rr_ratio"]), reverse=True)
    return overlap


def write_overlap(output_dir, rows):
    fields = [
        "watch_level",
        "mode_count",
        "code",
        "name",
        "exchange",
        "modes",
        "levels",
        "avg_score",
        "min_score",
        "entry_price",
        "nearest_target_price",
        "nearest_stop_price",
        "min_net_profit",
        "min_rr_ratio",
        "return_pct",
        "volume",
        "reasons",
    ]
    write_csv(output_dir / "overlap_candidates.csv", rows, fields)
    (output_dir / "overlap_candidates.json").write_text(
        json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def write_report(path, run_time, simulation, settings, groups, overlap_rows, excluded_count):
    lines = [
        f"# Stock Screen Report {run_time:%Y-%m-%d %H:%M:%S}",
        "",
        f"- Simulation: {simulation}",
        f"- Max price: {settings.max_price}",
        f"- Max one-lot amount: {settings.max_lot_amount:.0f}",
        f"- Min risk/reward: {settings.min_rr_ratio}",
        f"- Excluded rows: {excluded_count}",
        "",
        "## Overlap watchlist",
        "",
        "- core_watch: appears in daytrade, overnight, and weekly lists",
        "- priority_watch: appears in two lists",
        "- This is a watchlist, not a buy signal.",
        "",
        "| Rank | Watch | Modes | Code | Name | Avg score | Entry | Near target | Near stop | Min net | Min RR | Return | Volume |",
        "|---:|---|---|---|---|---:|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for index, row in enumerate(overlap_rows[:30], start=1):
        lines.append(
            f"| {index} | {row['watch_level']} | {row['modes']} | {row['code']} | {row['name']} | "
            f"{row['avg_score']:.1f} | {row['entry_price']:.2f} | {row['nearest_target_price']:.2f} | "
            f"{row['nearest_stop_price']:.2f} | {row['min_net_profit']:.0f} | {row['min_rr_ratio']:.2f} | "
            f"{row['return_pct']:.2f}% | {row['volume']} |"
        )
    if not overlap_rows:
        lines.append("| - | No overlap candidates | - | - | - | - | - | - | - | - | - | - | - |")
    lines.append("")

    labels = {
        "pre_breakout": "Pre-breakout candidates",
        "daytrade": "Daytrade candidates",
        "overnight": "Overnight candidates",
        "weekly": "Weekly candidates",
    }
    for mode, rows in groups.items():
        a_rows = [row for row in rows if row.level == "A"]
        lines.extend(
            [
                f"## {labels[mode]}",
                "",
                f"- A candidates: {len(a_rows)}",
                f"- Total candidates: {len(rows)}",
                "",
                "| Rank | Code | Name | Level | Score | Entry | Target | Stop | Net target | RR | Return | Volume | Reasons |",
                "|---:|---|---|---|---:|---:|---:|---:|---:|---:|---:|---:|---|",
            ]
        )
        for index, row in enumerate(rows[:30], start=1):
            lines.append(
                f"| {index} | {row.code} | {row.name} | {row.level} | {row.score:.1f} | "
                f"{row.entry_price:.2f} | {row.target_price:.2f} | {row.stop_price:.2f} | "
                f"{row.net_profit_at_target:.0f} | {row.rr_ratio:.2f} | {row.return_pct:.2f}% | "
                f"{row.volume} | {row.reasons} |"
            )
        if not rows:
            lines.append("| - | - | No candidates | - | - | - | - | - | - | - | - | - | - |")
        lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def rank_snapshot_pool(rows, settings):
    ranked = []
    for row in rows:
        if common_exclusions(row, settings):
            continue
        scores = snapshot_base_scores(row)
        ranked.append((scores["liquidity"] + scores["momentum"] + scores["technical"], row))
    ranked.sort(key=lambda item: item[0], reverse=True)
    return [row for _, row in ranked[: settings.history_pool_size]]


def run_once(output_label=None):
    run_time = datetime.now()
    output_dir = DATA_DIR / run_time.strftime("%Y-%m-%d")
    output_dir.mkdir(parents=True, exist_ok=True)

    api, simulation = load_api()
    settings = load_settings()
    try:
        snapshot_rows, contracts_by_code = fetch_snapshots(api)
        history_pool = rank_snapshot_pool(snapshot_rows, settings)
        history_contracts = [contracts_by_code[row["code"]] for row in history_pool if row["code"] in contracts_by_code]
        history = fetch_history_features(api, history_contracts, settings.kbar_days)
    finally:
        safe_logout(api)

    snapshot_fields = [
        "ts",
        "code",
        "name",
        "exchange",
        "day_trade",
        "category",
        "reference",
        "limit_up",
        "limit_down",
        "open",
        "high",
        "low",
        "close",
        "buy_price",
        "sell_price",
        "volume",
        "total_amount",
    ]
    write_csv(output_dir / "snapshots.csv", snapshot_rows, snapshot_fields)

    groups = {"pre_breakout": [], "daytrade": [], "overnight": [], "weekly": []}
    excluded = []
    for row in snapshot_rows:
        for mode in groups:
            candidate, reasons = make_candidate(row, mode, settings, history.get(row["code"]))
            if candidate:
                groups[mode].append(candidate)
            elif reasons:
                excluded.append({"mode": mode, **row, "exclude_reasons": ",".join(reasons)})

    for mode in groups:
        groups[mode].sort(key=lambda item: item.score, reverse=True)
        write_candidates(output_dir, f"{mode}_andidates" if False else f"{mode}_candidates", groups[mode])

    overlap_rows = build_overlap_candidates(groups)
    write_overlap(output_dir, overlap_rows)
    write_candidates(output_dir, "candidates", groups["daytrade"])
    write_csv(output_dir / "excluded.csv", excluded, list(excluded[0].keys()) if excluded else ["mode", "code", "name", "exclude_reasons"])
    write_report(output_dir / "screening_report.md", run_time, simulation, settings, groups, overlap_rows, len(excluded))
    write_excel_report(output_dir, snapshot_rows, groups, overlap_rows, excluded)
    if output_label:
        write_report(
            output_dir / f"screening_report_{output_label}.md",
            run_time,
            simulation,
            settings,
            groups,
            overlap_rows,
            len(excluded),
        )
        write_excel_report(
            output_dir,
            snapshot_rows,
            groups,
            overlap_rows,
            excluded,
            output_name=f"stock_screen_report_{output_label}.xlsx",
        )
        copy_labeled_outputs(output_dir, output_label)

    print(f"output_dir={output_dir}")
    print(
        f"snapshots={len(snapshot_rows)} "
        f"pre_breakout={len(groups['pre_breakout'])} daytrade={len(groups['daytrade'])} overnight={len(groups['overnight'])} "
        f"weekly={len(groups['weekly'])} overlap={len(overlap_rows)} excluded={len(excluded)}"
    )
    print("[overlap]")
    for row in overlap_rows[:5]:
        print(
            f"{row['watch_level']} {row['code']} {row['name']} modes={row['modes']} "
            f"avg_score={row['avg_score']:.1f} min_rr={row['min_rr_ratio']:.2f}"
        )
    for mode in ("pre_breakout", "daytrade", "overnight", "weekly"):
        print(f"[{mode}]")
        for row in groups[mode][:5]:
            print(
                f"{row.level} {row.code} {row.name} score={row.score:.1f} "
                f"entry={row.entry_price:.2f} target={row.target_price:.2f} rr={row.rr_ratio:.2f}"
            )


def run_window():
    today = datetime.now().date()
    start = datetime.combine(today, dtime(8, 30))
    end = datetime.combine(today, dtime(9, 10))
    now = datetime.now()
    if now < start:
        time.sleep((start - now).total_seconds())
    while datetime.now() < end:
        interval = 30 if datetime.now().time() < dtime(9, 0) else 5
        run_once()
        time.sleep(interval)
    run_once(output_label="0910")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--once", action="store_true", help="Run one current-time screen.")
    parser.add_argument("--window", action="store_true", help="Run 8:30-9:10 window collection.")
    parser.add_argument("--output-label", default=None, help="Also write labeled output copies, for example 0910 or 1000.")
    args = parser.parse_args()
    if args.window:
        run_window()
    else:
        run_once(output_label=args.output_label)


if __name__ == "__main__":
    main()
