import argparse
import csv
import json
import os
import re
import shutil
import sys
import subprocess
import time
import urllib.request
from dataclasses import asdict, dataclass
from datetime import datetime, timedelta, time as dtime, date
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
CODE_RE = re.compile(r"^[1-9][0-9]{3}$")

HEADER_MAP = {
    "watch_level": "觀察等級",
    "mode_count": "上榜模式數",
    "code": "股票代號",
    "name": "股票名稱",
    "exchange": "市場",
    "modes": "上榜模式",
    "levels": "各模式等級",
    "avg_score": "平均分數",
    "min_score": "最低分數",
    "entry_price": "參考進場價",
    "nearest_target_price": "最近停利價",
    "nearest_stop_price": "最近停損價",
    "min_net_profit": "最低預估淨利",
    "min_rr_ratio": "最低風報比",
    "return_pct": "漲跌幅%",
    "volume": "成交量",
    "reasons": "入選原因",
    "mode": "模式",
    "level": "等級",
    "score": "分數",
    "target_price": "停利價",
    "stop_price": "停損價",
    "lot_amount": "單張金額",
    "net_profit_at_target": "停利預估淨利",
    "loss_at_stop": "停損預估虧損",
    "rr_ratio": "風報比",
    "gap_pct": "開盤跳空%",
    "fade_pct": "開盤後漲跌%",
    "spread_pct": "買賣價差%",
    "total_amount": "成交金額",
    "liquidity_score": "流動性分數",
    "momentum_score": "動能分數",
    "technical_score": "技術分數",
    "theme_score": "題材分數",
    "risk_score": "風險分數",
    "exclude_reasons": "排除原因",
    "ts": "資料時間",
    "day_trade": "可否當沖",
    "category": "產業代碼",
    "reference": "參考價",
    "limit_up": "漲停價",
    "limit_down": "跌停價",
    "open": "開盤價",
    "high": "最高價",
    "low": "最低價",
    "close": "成交/收盤價",
    "buy_price": "買一價",
    "sell_price": "賣一價",
}

MODE_MAP = {"daytrade": "當沖", "overnight": "隔日沖", "weekly": "週沖", "pre_breakout": "起漲點A級", "pre_breakout_watch": "起漲觀察B級", "stable": "近期穩健"}
SHEET_NOTES = {
    "重疊名單": "同時出現在多個模式的優先觀察池；用來排序觀察，不是直接買進訊號。",
    "起漲點A級": "嚴格確認型：開盤站上 MA5/MA10、守開盤、有量、未過熱，寧可少量不硬湊。",
    "起漲觀察B級": "早期觀察池：包含所有A級；另納入收盤/目前價突破 MA5 且接近 MA10 的準備起跑股，只做觀察。",
    "近期穩健": "找近 5-20 日趨勢穩定、均線排列較佳、波動較低且未明顯過熱的標的。",
    "當沖": "早盤短線：重視流動性、價量動能、買賣價差、守開盤與接近日高；只適合當日快速處理。",
    "隔日沖": "為隔日準備：重視日線趨勢、MA5/MA10、量能延續與收盤位置；14:00 收盤後版不因缺失買賣價差直接排除。",
    "週沖": "數日到一週觀察：重視 MA5/MA10/MA20 排列、近五日價量與趨勢延續，但避免已過熱。",
    "排除清單": "符合任一風險或門檻不足條件即排除；先看排除原因再決定是否人工追蹤。",
    "原始快照": "Shioaji 當次取得的原始報價快照，供回查價格、成交量、價差與可否當沖。",
}
WATCH_MAP = {
    "core_watch_strong": "核心觀察-強",
    "core_watch_with_breakout": "核心觀察-含起漲",
    "core_watch": "核心觀察",
    "priority_watch": "優先觀察",
}

DEFAULT_THEME_GROUPS = {
    "memory": ["2344", "2408", "2337", "3006", "3260", "8299"],
    "panel": ["2409", "3481", "6116", "4935", "5371"],
    "power_energy": ["1504", "1513", "1514", "1605", "6282", "6806", "1519", "8996"],
    "aerospace_defense": ["2634", "8222", "8033", "4572", "3004"],
    "semiconductor_equipment": ["3167", "3583", "6187", "6438", "6510", "8028", "3131"],
    "testing_packaging": ["2449", "6239", "3711", "3265", "6257", "2329"],
    "networking": ["2345", "2419", "5388", "3596", "4906", "6285"],
    "power_supply": ["2308", "2327", "2353", "3017", "3653", "6415"],
    "robotics_automation": ["1504", "2049", "2359", "2464", "3379", "4576"],
    "finance": ["2881", "2882", "2883", "2884", "2885", "2886", "2890", "2891", "2892"],
    "construction": ["2501", "2504", "2511", "2542", "2548", "5534", "6177"],
}

THEME_GROUPS = dict(DEFAULT_THEME_GROUPS)
THEME_META = {name: {"label": name, "hot": False, "bonus": 0} for name in THEME_GROUPS}
CODE_THEMES = {}
CORPORATE_ACTION_RISKS = {}


def normalize_theme_groups(raw):
    groups = {}
    meta = {}
    if not isinstance(raw, dict):
        return groups, meta
    themes = raw.get("themes", raw)
    for theme_name, value in themes.items():
        if isinstance(value, dict):
            codes = value.get("codes", [])
            label = value.get("label", theme_name)
            hot = bool(value.get("hot", False))
            bonus = int(value.get("bonus", 0) or 0)
        else:
            codes = value
            label = theme_name
            hot = False
            bonus = 0
        clean_codes = []
        for code in codes:
            text = str(code).strip()
            if CODE_RE.match(text):
                clean_codes.append(text)
        if clean_codes:
            groups[str(theme_name)] = clean_codes
            meta[str(theme_name)] = {"label": str(label), "hot": hot, "bonus": bonus}
    return groups, meta


def build_code_themes(theme_groups):
    code_themes = {}
    for theme_name, codes in theme_groups.items():
        for theme_code in codes:
            code_themes.setdefault(theme_code, []).append(theme_name)
    return code_themes


def load_theme_groups():
    groups = dict(DEFAULT_THEME_GROUPS)
    meta = {name: {"label": name, "hot": False, "bonus": 0} for name in groups}
    path = Path(os.getenv("STOCK_THEME_FILE", BASE_DIR / "themes_recent.json"))
    if path.exists():
        try:
            configured_groups, configured_meta = normalize_theme_groups(
                json.loads(path.read_text(encoding="utf-8-sig"))
            )
            groups.update(configured_groups)
            meta.update(configured_meta)
        except Exception as exc:
            print(f"theme config skipped: {path} ({exc})", file=sys.stderr)
    return groups, meta, build_code_themes(groups)


def refresh_theme_groups():
    global THEME_GROUPS, THEME_META, CODE_THEMES
    THEME_GROUPS, THEME_META, CODE_THEMES = load_theme_groups()


def parse_iso_date(value):
    if not value:
        return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    match = re.match(r"^(\d{2,3})年(\d{1,2})月(\d{1,2})日$", text)
    if match:
        roc_year, month, day = map(int, match.groups())
        return date(roc_year + 1911, month, day)
    match = re.match(r"^(\d{2,3})/(\d{1,2})/(\d{1,2})$", text)
    if match:
        roc_year, month, day = map(int, match.groups())
        return date(roc_year + 1911, month, day)
    return None


def http_json(url):
    request = urllib.request.Request(url, headers={"User-Agent": "stock-screen/1.0"})
    try:
        with urllib.request.urlopen(request, timeout=20) as response:
            return json.loads(response.read().decode("utf-8-sig"))
    except Exception:
        command = "$ProgressPreference='SilentlyContinue'; (Invoke-WebRequest -Uri $env:STOCK_SCREEN_PUBLIC_URL -UseBasicParsing -TimeoutSec 20).Content"
        env = os.environ.copy()
        env["STOCK_SCREEN_PUBLIC_URL"] = url
        result = subprocess.run(
            ["powershell.exe", "-NoProfile", "-Command", command],
            capture_output=True,
            text=True,
            encoding="utf-8-sig",
            timeout=30,
            check=True,
            env=env,
        )
        return json.loads(result.stdout)


def check_twse_trading_day_after_close(run_date):
    """Return (is_trading_day, reason) for after-close scheduled runs.

    This is intentionally conservative: if the public TWSE check is unreachable,
    keep running so a temporary network issue does not make us miss a trading day.
    """
    if run_date.weekday() >= 5:
        return False, "weekend"
    url = f"https://www.twse.com.tw/rwd/zh/afterTrading/MI_INDEX?response=json&date={run_date:%Y%m%d}&type=IND"
    try:
        payload = http_json(url)
    except Exception as exc:
        return True, f"trading_day_check_failed:{exc}"

    stat = str(payload.get("stat", "")).strip()
    has_market_data = any(payload.get(key) for key in ("data9", "data8", "data7", "tables"))
    if has_market_data:
        return True, "twse_market_data_available"
    if any(token in stat for token in ("很抱歉", "查無", "無資料", "no data", "No data")):
        return False, f"twse_no_market_data:{stat}"
    return True, f"twse_status_unknown:{stat or 'empty'}"


def write_non_trading_day_skip(output_dir, run_time, output_label, reason):
    label = output_label or "run"
    path = output_dir / f"market_closed_{label}.txt"
    path.write_text(
        "\n".join(
            [
                f"timestamp={run_time.isoformat(timespec='seconds')}",
                f"date={run_time.date().isoformat()}",
                f"label={label}",
                f"reason={reason}",
                "status=skipped_non_trading_day",
            ]
        ),
        encoding="utf-8",
    )

def fetch_twse_corporate_actions(run_date):
    url = f"https://www.twse.com.tw/rwd/zh/exRight/TWT48U?response=json&date={run_date:%Y%m%d}"
    payload = http_json(url)
    fields = payload.get("fields", [])
    entries = []
    for row in payload.get("data", []):
        item = dict(zip(fields, row))
        code = str(item.get("股票代號", "")).strip()
        event_date = parse_iso_date(item.get("除權除息日期"))
        if CODE_RE.match(code) and event_date:
            entries.append(
                {
                    "code": code,
                    "name": str(item.get("名稱", "")).strip(),
                    "kind": str(item.get("除權息", "ex_right_dividend")).strip(),
                    "ex_date": event_date.isoformat(),
                    "source": "twse",
                }
            )
    return entries


def fetch_tpex_corporate_actions(run_date):
    url = "https://www.tpex.org.tw/web/stock/exright/preAnnounce/prepost_result.php?l=zh-tw&o=json"
    payload = http_json(url)
    entries = []
    for table in payload.get("tables", []):
        fields = table.get("fields", [])
        for row in table.get("data", []):
            item = dict(zip(fields, row))
            code = str(item.get("代號", "")).strip()
            event_date = parse_iso_date(item.get("除權息日期"))
            if CODE_RE.match(code) and event_date:
                entries.append(
                    {
                        "code": code,
                        "name": str(item.get("名稱", "")).strip(),
                        "kind": str(item.get("除權息", "ex_right_dividend")).strip(),
                        "ex_date": event_date.isoformat(),
                        "source": "tpex",
                    }
                )
    return entries


def load_manual_corporate_action_entries(settings):
    path = Path(os.getenv("CORPORATE_ACTION_FILE", BASE_DIR / "corporate_actions_watchlist.json"))
    if not path.exists():
        return []
    try:
        payload = json.loads(path.read_text(encoding="utf-8-sig"))
    except Exception as exc:
        print(f"corporate action config skipped: {path} ({exc})", file=sys.stderr)
        return []
    entries = payload.get("entries", payload if isinstance(payload, list) else [])
    clean = []
    for item in entries:
        if not isinstance(item, dict) or item.get("enabled", True) is False:
            continue
        code = str(item.get("code", "")).strip()
        event_date = parse_iso_date(item.get("ex_date") or item.get("date"))
        if CODE_RE.match(code) and event_date:
            clean.append(
                {
                    "code": code,
                    "name": str(item.get("name", "")).strip(),
                    "kind": str(item.get("kind", "manual_corporate_action")),
                    "ex_date": event_date.isoformat(),
                    "source": "manual",
                    "before_days": item.get("before_days"),
                    "after_days": item.get("after_days"),
                }
            )
    return clean


def corporate_action_windows(settings):
    return {
        "daytrade": (settings.daytrade_corporate_action_before_days, settings.daytrade_corporate_action_after_days),
        "overnight": (settings.overnight_corporate_action_before_days, settings.overnight_corporate_action_after_days),
        "weekly": (settings.weekly_corporate_action_before_days, settings.weekly_corporate_action_after_days),
        "pre_breakout": (settings.pre_breakout_corporate_action_before_days, settings.pre_breakout_corporate_action_after_days),
        "pre_breakout_watch": (settings.pre_breakout_corporate_action_before_days, settings.pre_breakout_corporate_action_after_days),
    }


def load_corporate_action_risks(settings, run_date=None):
    if not settings.corporate_action_filter_enabled:
        return {}
    run_date = run_date or datetime.now().date()
    entries = load_manual_corporate_action_entries(settings)
    if settings.corporate_action_auto_enabled:
        for source_name, fetcher in (("twse", fetch_twse_corporate_actions), ("tpex", fetch_tpex_corporate_actions)):
            try:
                entries.extend(fetcher(run_date))
            except Exception as exc:
                print(f"corporate action auto source skipped: {source_name} ({exc})", file=sys.stderr)
    risks = {mode: {} for mode in corporate_action_windows(settings)}
    for item in entries:
        code = item["code"]
        event_date = parse_iso_date(item.get("ex_date"))
        if not event_date:
            continue
        source = item.get("source", "unknown")
        kind = item.get("kind", "corporate_action")
        for mode, (before_days, after_days) in corporate_action_windows(settings).items():
            before = int(item.get("before_days") or before_days)
            after = int(item.get("after_days") or after_days)
            if event_date - timedelta(days=before) <= run_date <= event_date + timedelta(days=after):
                risks[mode][code] = f"corporate_action_window:{source}:{kind}:{event_date.isoformat()}"
    return risks


def get_corporate_action_reason(code, mode=None):
    if mode:
        return CORPORATE_ACTION_RISKS.get(mode, {}).get(code)
    for mode_risks in CORPORATE_ACTION_RISKS.values():
        if code in mode_risks:
            return mode_risks[code]
    return None


def refresh_corporate_action_risks(settings, run_date=None):
    global CORPORATE_ACTION_RISKS
    CORPORATE_ACTION_RISKS = load_corporate_action_risks(settings, run_date)


refresh_theme_groups()


@dataclass
class Settings:
    max_price: float = 200.0
    max_lot_amount: float = 200_000.0
    min_volume: int = 1000
    max_spread_pct: float = 0.6
    commission_rate: float = 0.001425
    commission_discount: float = 0.6
    min_commission: float = 20.0
    intraday_tax_rate: float = 0.0015
    regular_tax_rate: float = 0.003
    daytrade_target_pct: float = 2.5
    daytrade_stop_pct: float = 0.6
    daytrade_min_net_profit: float = 800.0
    overnight_target_pct: float = 5.0
    overnight_stop_pct: float = 1.5
    overnight_min_net_profit: float = 1500.0
    weekly_target_pct: float = 8.0
    weekly_stop_pct: float = 3.0
    weekly_min_net_profit: float = 2500.0
    stable_target_pct: float = 6.0
    stable_stop_pct: float = 2.0
    stable_min_net_profit: float = 1200.0
    pre_breakout_target_pct: float = 6.5
    pre_breakout_stop_pct: float = 2.0
    pre_breakout_min_net_profit: float = 1200.0
    pre_breakout_min_gap_pct: float = 0.5
    pre_breakout_max_gap_pct: float = 3.5
    pre_breakout_max_return_pct: float = 4.0
    pre_breakout_max_ma5_distance_pct: float = 6.0
    pre_breakout_max_ma20_distance_pct: float = 12.0
    pre_breakout_max_above_prev_high_pct: float = 2.0
    pre_breakout_max_20d_return_pct: float = 18.0
    min_rr_ratio: float = 2.0
    history_pool_size: int = 80
    kbar_days: int = 70
    corporate_action_filter_enabled: bool = True
    corporate_action_before_days: int = 3
    corporate_action_after_days: int = 5
    corporate_action_auto_enabled: bool = True
    daytrade_corporate_action_before_days: int = 1
    daytrade_corporate_action_after_days: int = 2
    overnight_corporate_action_before_days: int = 2
    overnight_corporate_action_after_days: int = 5
    weekly_corporate_action_before_days: int = 2
    weekly_corporate_action_after_days: int = 5
    pre_breakout_corporate_action_before_days: int = 3
    pre_breakout_corporate_action_after_days: int = 5


@dataclass
class Candidate:
    mode: str
    code: str
    name: str
    exchange: str
    level: str
    score: float
    entry_price: float
    target_price: float
    stop_price: float
    lot_amount: float
    net_profit_at_target: float
    loss_at_stop: float
    rr_ratio: float
    return_pct: float
    gap_pct: float
    fade_pct: float
    spread_pct: float
    volume: int
    total_amount: float
    liquidity_score: int
    momentum_score: int
    technical_score: int
    theme_score: int
    risk_score: int
    reasons: str


def env_float(name, default):
    value = os.getenv(name)
    if value in (None, ""):
        return default
    return float(value)


def env_int(name, default):
    value = os.getenv(name)
    if value in (None, ""):
        return default
    return int(value)


def env_bool(name, default):
    value = os.getenv(name)
    if value in (None, ""):
        return default
    return str(value).strip().lower() not in ("0", "false", "no", "off")


def load_settings():
    return Settings(
        max_price=env_float("STOCK_MAX_PRICE", Settings.max_price),
        max_lot_amount=env_float("STOCK_MAX_LOT_AMOUNT", Settings.max_lot_amount),
        min_volume=env_int("STOCK_MIN_VOLUME", Settings.min_volume),
        max_spread_pct=env_float("STOCK_MAX_SPREAD_PCT", Settings.max_spread_pct),
        commission_rate=env_float("STOCK_COMMISSION_RATE", Settings.commission_rate),
        commission_discount=env_float("STOCK_COMMISSION_DISCOUNT", Settings.commission_discount),
        min_commission=env_float("STOCK_MIN_COMMISSION", Settings.min_commission),
        intraday_tax_rate=env_float("STOCK_INTRADAY_TAX_RATE", Settings.intraday_tax_rate),
        regular_tax_rate=env_float("STOCK_REGULAR_TAX_RATE", Settings.regular_tax_rate),
        daytrade_target_pct=env_float("DAYTRADE_TARGET_PCT", Settings.daytrade_target_pct),
        daytrade_stop_pct=env_float("DAYTRADE_STOP_PCT", Settings.daytrade_stop_pct),
        daytrade_min_net_profit=env_float("DAYTRADE_MIN_NET_PROFIT", Settings.daytrade_min_net_profit),
        overnight_target_pct=env_float("OVERNIGHT_TARGET_PCT", Settings.overnight_target_pct),
        overnight_stop_pct=env_float("OVERNIGHT_STOP_PCT", Settings.overnight_stop_pct),
        overnight_min_net_profit=env_float("OVERNIGHT_MIN_NET_PROFIT", Settings.overnight_min_net_profit),
        weekly_target_pct=env_float("WEEKLY_TARGET_PCT", Settings.weekly_target_pct),
        weekly_stop_pct=env_float("WEEKLY_STOP_PCT", Settings.weekly_stop_pct),
        weekly_min_net_profit=env_float("WEEKLY_MIN_NET_PROFIT", Settings.weekly_min_net_profit),
        stable_target_pct=env_float("STABLE_TARGET_PCT", Settings.stable_target_pct),
        stable_stop_pct=env_float("STABLE_STOP_PCT", Settings.stable_stop_pct),
        stable_min_net_profit=env_float("STABLE_MIN_NET_PROFIT", Settings.stable_min_net_profit),
        pre_breakout_target_pct=env_float("PRE_BREAKOUT_TARGET_PCT", Settings.pre_breakout_target_pct),
        pre_breakout_stop_pct=env_float("PRE_BREAKOUT_STOP_PCT", Settings.pre_breakout_stop_pct),
        pre_breakout_min_net_profit=env_float("PRE_BREAKOUT_MIN_NET_PROFIT", Settings.pre_breakout_min_net_profit),
        pre_breakout_min_gap_pct=env_float("PRE_BREAKOUT_MIN_GAP_PCT", Settings.pre_breakout_min_gap_pct),
        pre_breakout_max_gap_pct=env_float("PRE_BREAKOUT_MAX_GAP_PCT", Settings.pre_breakout_max_gap_pct),
        pre_breakout_max_return_pct=env_float("PRE_BREAKOUT_MAX_RETURN_PCT", Settings.pre_breakout_max_return_pct),
        pre_breakout_max_ma5_distance_pct=env_float("PRE_BREAKOUT_MAX_MA5_DISTANCE_PCT", Settings.pre_breakout_max_ma5_distance_pct),
        pre_breakout_max_ma20_distance_pct=env_float("PRE_BREAKOUT_MAX_MA20_DISTANCE_PCT", Settings.pre_breakout_max_ma20_distance_pct),
        pre_breakout_max_above_prev_high_pct=env_float("PRE_BREAKOUT_MAX_ABOVE_PREV_HIGH_PCT", Settings.pre_breakout_max_above_prev_high_pct),
        pre_breakout_max_20d_return_pct=env_float("PRE_BREAKOUT_MAX_20D_RETURN_PCT", Settings.pre_breakout_max_20d_return_pct),
        min_rr_ratio=env_float("STOCK_MIN_RR_RATIO", Settings.min_rr_ratio),
        history_pool_size=env_int("STOCK_HISTORY_POOL_SIZE", Settings.history_pool_size),
        kbar_days=env_int("STOCK_KBAR_DAYS", Settings.kbar_days),
        corporate_action_filter_enabled=env_bool("CORPORATE_ACTION_FILTER_ENABLED", Settings.corporate_action_filter_enabled),
        corporate_action_before_days=env_int("CORPORATE_ACTION_BEFORE_DAYS", Settings.corporate_action_before_days),
        corporate_action_after_days=env_int("CORPORATE_ACTION_AFTER_DAYS", Settings.corporate_action_after_days),
        corporate_action_auto_enabled=env_bool("CORPORATE_ACTION_AUTO_ENABLED", Settings.corporate_action_auto_enabled),
        daytrade_corporate_action_before_days=env_int("DAYTRADE_CORPORATE_ACTION_BEFORE_DAYS", Settings.daytrade_corporate_action_before_days),
        daytrade_corporate_action_after_days=env_int("DAYTRADE_CORPORATE_ACTION_AFTER_DAYS", Settings.daytrade_corporate_action_after_days),
        overnight_corporate_action_before_days=env_int("OVERNIGHT_CORPORATE_ACTION_BEFORE_DAYS", Settings.overnight_corporate_action_before_days),
        overnight_corporate_action_after_days=env_int("OVERNIGHT_CORPORATE_ACTION_AFTER_DAYS", Settings.overnight_corporate_action_after_days),
        weekly_corporate_action_before_days=env_int("WEEKLY_CORPORATE_ACTION_BEFORE_DAYS", Settings.weekly_corporate_action_before_days),
        weekly_corporate_action_after_days=env_int("WEEKLY_CORPORATE_ACTION_AFTER_DAYS", Settings.weekly_corporate_action_after_days),
        pre_breakout_corporate_action_before_days=env_int("PRE_BREAKOUT_CORPORATE_ACTION_BEFORE_DAYS", Settings.pre_breakout_corporate_action_before_days),
        pre_breakout_corporate_action_after_days=env_int("PRE_BREAKOUT_CORPORATE_ACTION_AFTER_DAYS", Settings.pre_breakout_corporate_action_after_days),
    )


def env_required(name):
    value = os.getenv(name)
    if not value:
        raise RuntimeError(f"Missing required environment variable: {name}")
    return value


def to_float(value, default=0.0):
    try:
        if value is None:
            return default
        return float(value)
    except Exception:
        return default


def to_exchange(value):
    text = str(value)
    if "." in text:
        text = text.rsplit(".", 1)[-1]
    return text.replace("Exchange.", "")


def load_api():
    load_dotenv(BASE_DIR / ".env", encoding="utf-8-sig")
    import shioaji as sj

    simulation = os.getenv("SHIOAJI_SIMULATION", "true").lower() != "false"
    api = sj.Shioaji(simulation=simulation)
    api.login(api_key=env_required("SHIOAJI_API_KEY"), secret_key=env_required("SHIOAJI_SECRET_KEY"))
    return api, simulation


def safe_logout(api):
    try:
        api.logout()
    except Exception as exc:
        print(f"logout skipped: {exc}", file=sys.stderr)


def iter_stock_contracts(api):
    for group_name in ("TSE", "OTC"):
        group = getattr(api.Contracts.Stocks, group_name)
        for contract in group:
            code = getattr(contract, "code", "")
            if CODE_RE.match(code):
                yield contract


def chunks(items, size):
    for index in range(0, len(items), size):
        yield items[index : index + size]


def snapshot_to_dict(snapshot, contract):
    return {
        "code": getattr(contract, "code", ""),
        "name": getattr(contract, "name", ""),
        "exchange": to_exchange(getattr(contract, "exchange", "")),
        "day_trade": str(getattr(contract, "day_trade", "")),
        "category": getattr(contract, "category", ""),
        "reference": to_float(getattr(contract, "reference", 0)),
        "limit_up": to_float(getattr(contract, "limit_up", 0)),
        "limit_down": to_float(getattr(contract, "limit_down", 0)),
        "open": to_float(getattr(snapshot, "open", 0)),
        "high": to_float(getattr(snapshot, "high", 0)),
        "low": to_float(getattr(snapshot, "low", 0)),
        "close": to_float(getattr(snapshot, "close", 0)),
        "buy_price": to_float(getattr(snapshot, "buy_price", 0)),
        "sell_price": to_float(getattr(snapshot, "sell_price", 0)),
        "volume": int(to_float(getattr(snapshot, "total_volume", 0))),
        "total_amount": to_float(getattr(snapshot, "total_amount", 0)),
        "ts": datetime.now().isoformat(timespec="seconds"),
    }


def fetch_snapshots(api):
    contracts = list(iter_stock_contracts(api))
    by_code = {contract.code: contract for contract in contracts}
    rows = []
    for batch in chunks(contracts, 200):
        try:
            snapshots = api.snapshots(batch)
        except Exception as exc:
            print(f"snapshot batch failed: {exc}", file=sys.stderr)
            continue
        for snapshot in snapshots:
            contract = by_code.get(str(getattr(snapshot, "code", "")))
            if contract:
                rows.append(snapshot_to_dict(snapshot, contract))
        time.sleep(0.1)
    return rows, by_code


def fee(amount, settings):
    return max(settings.min_commission, amount * settings.commission_rate * settings.commission_discount)


def trade_math(entry, target, stop, mode, settings):
    shares = 1000
    tax_rate = settings.intraday_tax_rate if mode == "daytrade" else settings.regular_tax_rate
    buy_amount = entry * shares
    target_amount = target * shares
    stop_amount = stop * shares

    buy_fee = fee(buy_amount, settings)
    sell_fee_target = fee(target_amount, settings)
    sell_fee_stop = fee(stop_amount, settings)
    tax_target = target_amount * tax_rate
    tax_stop = stop_amount * tax_rate

    net_profit = target_amount - buy_amount - buy_fee - sell_fee_target - tax_target
    stop_loss = buy_amount - stop_amount + buy_fee + sell_fee_stop + tax_stop
    rr = net_profit / stop_loss if stop_loss > 0 else 0
    return round(net_profit, 0), round(stop_loss, 0), round(rr, 2)


def common_exclusions(row, settings, mode=None, after_close=False):
    reasons = []
    close = row["close"]
    reference = row["reference"]
    buy_price = row["buy_price"]
    sell_price = row["sell_price"]
    has_live_spread = bool(close and buy_price and sell_price)
    spread_pct = (sell_price - buy_price) / close * 100 if has_live_spread else 999

    if close <= 0 or reference <= 0:
        reasons.append("invalid_quote")
    if row["volume"] < settings.min_volume:
        reasons.append("volume_too_low")
    if spread_pct > settings.max_spread_pct and not (after_close and not has_live_spread):
        reasons.append("spread_too_wide")
    if close > settings.max_price:
        reasons.append("price_above_limit")
    if close * 1000 > settings.max_lot_amount:
        reasons.append("one_lot_amount_above_limit")
    corporate_action_risk = get_corporate_action_reason(row["code"], mode)
    if corporate_action_risk:
        reasons.append(corporate_action_risk)
    return reasons


def mode_trade_plan(row, mode, settings):
    entry = row["close"]
    if mode == "daytrade":
        target_pct = settings.daytrade_target_pct
        stop_pct = settings.daytrade_stop_pct
        min_net = settings.daytrade_min_net_profit
    elif mode == "overnight":
        target_pct = settings.overnight_target_pct
        stop_pct = settings.overnight_stop_pct
        min_net = settings.overnight_min_net_profit
    elif mode == "weekly":
        target_pct = settings.weekly_target_pct
        stop_pct = settings.weekly_stop_pct
        min_net = settings.weekly_min_net_profit
    elif mode == "stable":
        target_pct = settings.stable_target_pct
        stop_pct = settings.stable_stop_pct
        min_net = settings.stable_min_net_profit
    else:
        target_pct = settings.pre_breakout_target_pct
        stop_pct = settings.pre_breakout_stop_pct
        min_net = settings.pre_breakout_min_net_profit

    target = round(entry * (1 + target_pct / 100), 2)
    stop = round(entry * (1 - stop_pct / 100), 2)
    net_profit, stop_loss, rr = trade_math(entry, target, stop, mode, settings)
    reasons = []
    if net_profit < min_net:
        reasons.append("net_profit_below_min")
    if rr < settings.min_rr_ratio:
        reasons.append("rr_below_min")
    return target, stop, net_profit, stop_loss, rr, reasons


def snapshot_base_scores(row):
    close = row["close"]
    open_price = row["open"]
    high = row["high"]
    low = row["low"]
    reference = row["reference"]
    buy_price = row["buy_price"]
    sell_price = row["sell_price"]
    volume = row["volume"]

    return_pct = (close - reference) / reference * 100 if reference else 0
    gap_pct = (open_price - reference) / reference * 100 if reference else 0
    fade_pct = (close - open_price) / open_price * 100 if open_price else 0
    has_live_spread = bool(close and buy_price and sell_price)
    spread_pct = (sell_price - buy_price) / close * 100 if has_live_spread else 999
    range_pct = (high - low) / reference * 100 if reference else 0

    liquidity = 0
    liquidity += 8 if volume >= 10000 else 6 if volume >= 5000 else 4 if volume >= 2000 else 2
    liquidity += 6 if row["total_amount"] >= 500_000_000 else 4 if row["total_amount"] >= 150_000_000 else 2
    liquidity += 4 if spread_pct <= 0.2 else 3 if spread_pct <= 0.4 else 1

    momentum = 0
    momentum += 5 if gap_pct >= 2 else 3 if gap_pct >= 1 else 1 if gap_pct > 0 else 0
    momentum += 5 if return_pct >= 4 else 4 if return_pct >= 3 else 3 if return_pct >= 2 else 1 if return_pct > 0 else 0
    if gap_pct > 0 and fade_pct >= 0 and return_pct > 1:
        momentum += 3

    technical = 0
    if close > reference:
        technical += 3
    if close > open_price:
        technical += 2
    if 2 <= range_pct <= 6:
        technical += 2
    if fade_pct >= 0:
        technical += 2
    if spread_pct <= 0.2:
        technical += 1

    risk = 5
    if range_pct > 6:
        risk -= 2
    if spread_pct > 0.4:
        risk -= 2

    reasons = []
    if close > open_price:
        reasons.append("above_open")
    if high > low and (high - close) / (high - low) <= 0.25:
        reasons.append("near_day_high")
    if return_pct >= 4 and volume >= 3000:
        reasons.append("strong_momentum")

    return {
        "return_pct": return_pct,
        "gap_pct": gap_pct,
        "fade_pct": fade_pct,
        "spread_pct": spread_pct,
        "liquidity": liquidity,
        "momentum": min(15, momentum),
        "technical": min(10, technical),
        "risk": max(0, risk),
        "reasons": reasons,
    }


def build_theme_features(snapshot_rows):
    rows_by_code = {row["code"]: row for row in snapshot_rows}
    features = {}
    for theme_name, codes in THEME_GROUPS.items():
        members = [rows_by_code[code] for code in codes if code in rows_by_code]
        if not members:
            continue
        returns = []
        positive = 0
        strong = 0
        above_open = 0
        high_volume = 0
        for row in members:
            reference = row["reference"]
            close = row["close"]
            return_pct = (close - reference) / reference * 100 if reference else 0
            returns.append(return_pct)
            if return_pct > 0:
                positive += 1
            if return_pct >= 2:
                strong += 1
            if close >= row["open"] and close > reference:
                above_open += 1
            if row["volume"] >= 5000 or row["total_amount"] >= 150_000_000:
                high_volume += 1
        theme_data = {
            "member_count": len(members),
            "positive_count": positive,
            "strong_count": strong,
            "above_open_count": above_open,
            "high_volume_count": high_volume,
            "avg_return_pct": avg(returns),
        }
        features[theme_name] = theme_data
    return features


def theme_score_for(row, theme_features):
    themes = CODE_THEMES.get(row["code"], [])
    if not themes:
        return 0, []
    best_score = 0
    best_tags = []
    for theme_name in themes:
        data = theme_features.get(theme_name)
        if not data:
            continue
        score = 0
        meta = THEME_META.get(theme_name, {})
        tags = [f"theme:{theme_name}"]
        if meta.get("hot") or meta.get("bonus", 0) > 0:
            hot_bonus = min(8, max(3, int(meta.get("bonus", 0) or 0)))
            score += hot_bonus
            tags.append("hot_theme")
        if data["positive_count"] >= 3:
            score += 4
            tags.append("theme_breadth")
        if data["strong_count"] >= 2:
            score += 4
            tags.append("theme_strong_peers")
        if data["above_open_count"] >= 2:
            score += 3
            tags.append("theme_holds_open")
        if data["high_volume_count"] >= 2:
            score += 3
            tags.append("theme_volume")
        if data["avg_return_pct"] >= 1.5:
            score += 4
            tags.append("theme_avg_up")
        elif data["avg_return_pct"] >= 0.5:
            score += 2
            tags.append("theme_mild_up")
        score = min(22, score)
        if score > best_score:
            best_score = score
            best_tags = tags
    return best_score, best_tags


def make_candidate(row, mode, settings, history=None, theme_features=None, after_close=False):
    if mode == "daytrade" and "Yes" not in row["day_trade"]:
        return None, ["not_day_trade"]

    reasons = common_exclusions(row, settings, mode, after_close=after_close)
    target, stop, net_profit, stop_loss, rr, trade_reasons = mode_trade_plan(row, mode, settings)
    reasons.extend(trade_reasons)
    if reasons:
        return None, reasons

    scores = snapshot_base_scores(row)
    liquidity = scores["liquidity"]
    momentum = scores["momentum"]
    technical = scores["technical"]
    risk = scores["risk"]
    reason_tags = list(scores["reasons"])
    theme_score, theme_tags = theme_score_for(row, theme_features or {})
    reason_tags.extend(theme_tags)

    if mode == "overnight" and history:
        add, tags = overnight_history_score(history)
        technical += add
        reason_tags.extend(tags)
    elif mode == "weekly" and history:
        add, tags = weekly_history_score(history)
        technical += add
        reason_tags.extend(tags)
    elif mode == "stable":
        if not history:
            return None, ["not_enough_history"]
        add, tags, blockers = stable_history_score(history)
        if blockers:
            return None, blockers
        technical += add
        reason_tags.extend(tags)
    elif mode in ("pre_breakout", "pre_breakout_watch"):
        if not history:
            return None, ["not_enough_history"]
        add, tags, blockers = pre_breakout_history_score(row, history, settings, relaxed=(mode == "pre_breakout_watch"))
        if blockers:
            return None, blockers
        technical += add
        reason_tags.extend(tags)

    max_raw = 73 if mode == "daytrade" else 88 if mode in ("pre_breakout", "pre_breakout_watch") else 92 if mode == "stable" else 80
    raw = liquidity + momentum + min(20, technical) + risk + theme_score
    if mode == "daytrade":
        raw += 15 if row["close"] > row["open"] else 5
    elif mode in ("pre_breakout", "pre_breakout_watch"):
        raw += 12 if row["close"] >= row["open"] else 4 if mode == "pre_breakout_watch" else 0
        raw += min(8, theme_score // 2)
        if mode == "pre_breakout_watch":
            raw -= 8
    elif mode == "stable":
        raw += 10 if row["close"] >= row["open"] else 3
        raw += min(5, theme_score // 4)
    else:
        raw += 10 if scores["return_pct"] > 0 else 0

    score = min(100.0, round(raw / max_raw * 100, 1))
    level = "A" if score >= 80 and liquidity >= 14 and rr >= settings.min_rr_ratio else "B" if score >= 70 else "C"
    if level == "C":
        return None, ["score_below_b"]

    return Candidate(
        mode=mode,
        code=row["code"],
        name=row["name"],
        exchange=row["exchange"],
        level=level,
        score=score,
        entry_price=round(row["close"], 4),
        target_price=target,
        stop_price=stop,
        lot_amount=round(row["close"] * 1000, 0),
        net_profit_at_target=net_profit,
        loss_at_stop=stop_loss,
        rr_ratio=rr,
        return_pct=round(scores["return_pct"], 2),
        gap_pct=round(scores["gap_pct"], 2),
        fade_pct=round(scores["fade_pct"], 2),
        spread_pct=round(scores["spread_pct"], 3),
        volume=row["volume"],
        total_amount=round(row["total_amount"], 0),
        liquidity_score=liquidity,
        momentum_score=momentum,
        technical_score=min(20, technical),
        theme_score=theme_score,
        risk_score=risk,
        reasons=",".join(reason_tags[:8]),
    ), []


def daily_bars_from_kbars(kbars):
    grouped = {}
    for ts, open_, high, low, close, volume in zip(
        kbars.ts, kbars.Open, kbars.High, kbars.Low, kbars.Close, kbars.Volume
    ):
        day = datetime.fromtimestamp(ts / 1_000_000_000).date().isoformat()
        bucket = grouped.setdefault(
            day,
            {"date": day, "open": open_, "high": high, "low": low, "close": close, "volume": 0},
        )
        bucket["high"] = max(bucket["high"], high)
        bucket["low"] = min(bucket["low"], low)
        bucket["close"] = close
        bucket["volume"] += volume
    return [grouped[key] for key in sorted(grouped)]


def avg(values):
    return sum(values) / len(values) if values else 0


def fetch_history_features(api, contracts, days):
    end = datetime.now().date()
    start = end - timedelta(days=days)
    features = {}
    for contract in contracts:
        try:
            kbars = api.kbars(contract, start=start.isoformat(), end=end.isoformat())
            bars = daily_bars_from_kbars(kbars)
            if len(bars) >= 10:
                features[contract.code] = bars
        except Exception as exc:
            print(f"kbar failed {contract.code}: {exc}", file=sys.stderr)
            text = str(exc)
            if "SessionNotEstablished" in text or "NotReady" in text:
                break
        time.sleep(0.03)
    return features


def overnight_history_score(bars):
    tags = []
    closes = [bar["close"] for bar in bars]
    vols = [bar["volume"] for bar in bars]
    last = bars[-1]
    ma5 = avg(closes[-5:])
    ma10 = avg(closes[-10:])
    vol5 = avg(vols[-5:])
    score = 0
    if last["close"] > ma5:
        score += 4
        tags.append("above_ma5")
    if last["close"] > ma10:
        score += 3
        tags.append("above_ma10")
    if ma5 > ma10:
        score += 3
        tags.append("ma5_gt_ma10")
    if vol5 and last["volume"] > vol5 * 1.3:
        score += 4
        tags.append("volume_expansion")
    if last["high"] > last["low"] and (last["high"] - last["close"]) / (last["high"] - last["low"]) <= 0.25:
        score += 3
        tags.append("daily_close_near_high")
    return score, tags


def weekly_history_score(bars):
    tags = []
    closes = [bar["close"] for bar in bars]
    vols = [bar["volume"] for bar in bars]
    last = bars[-1]
    ma5 = avg(closes[-5:])
    ma10 = avg(closes[-10:])
    ma20 = avg(closes[-20:]) if len(closes) >= 20 else ma10
    vol20 = avg(vols[-20:]) if len(vols) >= 20 else avg(vols)
    score = 0
    if last["close"] > ma5 > ma10:
        score += 5
        tags.append("short_trend_up")
    if ma5 > ma10 > ma20:
        score += 5
        tags.append("ma_alignment")
    if closes[-1] > closes[-5]:
        score += 3
        tags.append("five_day_price_up")
    if vol20 and avg(vols[-5:]) > vol20 * 1.2:
        score += 4
        tags.append("five_day_volume_up")
    if last["close"] > ma20:
        score += 3
        tags.append("above_ma20")
    return score, tags


def stable_history_score(bars):
    if len(bars) < 20:
        return 0, [], ["not_enough_history"]
    tags = []
    blockers = []
    closes = [bar["close"] for bar in bars]
    lows = [bar["low"] for bar in bars]
    vols = [bar["volume"] for bar in bars]
    last_close = closes[-1]
    ma5 = avg(closes[-5:])
    ma10 = avg(closes[-10:])
    ma20 = avg(closes[-20:])
    ret5 = (last_close - closes[-6]) / closes[-6] * 100 if len(closes) >= 6 and closes[-6] else 0
    ret10 = (last_close - closes[-11]) / closes[-11] * 100 if len(closes) >= 11 and closes[-11] else 0
    ret20 = (last_close - closes[-21]) / closes[-21] * 100 if len(closes) >= 21 and closes[-21] else 0
    max20 = max(closes[-20:])
    drawdown20 = (last_close - max20) / max20 * 100 if max20 else 0
    ma20_dist = (last_close - ma20) / ma20 * 100 if ma20 else 999
    daily_returns = [((closes[i] - closes[i - 1]) / closes[i - 1] * 100) for i in range(1, len(closes)) if closes[i - 1]]
    recent_vol = 0
    if len(daily_returns) >= 10:
        mean_return = avg(daily_returns[-10:])
        recent_vol = (sum((value - mean_return) ** 2 for value in daily_returns[-10:]) / 10) ** 0.5
    higher_lows = sum(1 for index in range(max(0, len(lows) - 5), len(lows) - 1) if lows[index + 1] >= lows[index])
    vol5 = avg(vols[-5:])
    vol20 = avg(vols[-20:])

    if not (last_close > ma5 and last_close > ma10 and last_close > ma20):
        blockers.append("not_above_stable_ma")
    if ret20 > 22:
        blockers.append("stable_recent_runup_too_high")
    if ma20_dist > 10:
        blockers.append("stable_extended_from_ma20")
    if ret5 > 12:
        blockers.append("stable_short_term_overheat")
    if drawdown20 < -8:
        blockers.append("stable_drawdown_too_deep")

    score = 0
    if last_close > ma5 > ma10:
        score += 6
        tags.append("stable_short_trend")
    if ma5 >= ma10 >= ma20:
        score += 6
        tags.append("stable_ma_alignment")
    if 1 <= ret10 <= 10:
        score += 5
        tags.append("stable_10d_up_not_hot")
    if 3 <= ret20 <= 22:
        score += 5
        tags.append("stable_20d_up")
    if 0 <= ma20_dist <= 8:
        score += 4
        tags.append("stable_near_ma20")
    elif 8 < ma20_dist <= 10:
        score += 2
        tags.append("stable_ma20_not_too_far")
    if drawdown20 >= -4:
        score += 4
        tags.append("stable_shallow_pullback")
    if recent_vol and recent_vol <= 1.8:
        score += 4
        tags.append("stable_low_volatility")
    elif recent_vol and recent_vol <= 2.5:
        score += 2
        tags.append("stable_ok_volatility")
    if higher_lows >= 3:
        score += 3
        tags.append("stable_higher_lows")
    if vol20 and 0.6 <= vol5 / vol20 <= 1.8:
        score += 3
        tags.append("stable_volume_orderly")

    return score, tags, blockers


def calc_kd(bars, period=9):
    k_value = 50.0
    d_value = 50.0
    for index in range(len(bars)):
        start = max(0, index - period + 1)
        window = bars[start : index + 1]
        high = max(bar["high"] for bar in window)
        low = min(bar["low"] for bar in window)
        close = bars[index]["close"]
        rsv = 50.0 if high == low else (close - low) / (high - low) * 100
        k_value = k_value * 2 / 3 + rsv / 3
        d_value = d_value * 2 / 3 + k_value / 3
    return k_value, d_value


def pre_breakout_history_score(row, bars, settings, relaxed=False):
    if len(bars) < 20:
        return 0, [], ["not_enough_history"]

    # Shioaji kbars usually includes today's incomplete bar during trading hours.
    history = bars[:-1] if bars[-1]["date"] == datetime.now().date().isoformat() and len(bars) > 20 else bars
    if len(history) < 20:
        return 0, [], ["not_enough_history"]

    open_price = row["open"]
    close = row["close"]
    reference = row["reference"]
    volume = row["volume"]
    gap_pct = (open_price - reference) / reference * 100 if reference else 0
    return_pct = (close - reference) / reference * 100 if reference else 0
    intraday_range = row["high"] - row["low"]
    upper_fade_ratio = (row["high"] - close) / intraday_range if intraday_range > 0 else 0

    blockers = []
    min_gap = 0.0 if relaxed else settings.pre_breakout_min_gap_pct
    max_gap = 4.5 if relaxed else settings.pre_breakout_max_gap_pct
    max_return = 5.5 if relaxed else settings.pre_breakout_max_return_pct
    max_fade_ratio = 0.70 if relaxed else 0.55
    if gap_pct < min_gap:
        blockers.append("gap_too_small")
    if gap_pct > max_gap:
        blockers.append("gap_too_large")
    if return_pct > max_return:
        blockers.append("already_ran_intraday")
    if upper_fade_ratio > max_fade_ratio and row["high"] > open_price:
        blockers.append("intraday_spike_fade")

    closes = [bar["close"] for bar in history]
    vols = [bar["volume"] for bar in history]
    prev = history[-1]
    ma5 = avg(closes[-5:])
    ma10 = avg(closes[-10:])
    ma20 = avg(closes[-20:])
    ma60 = avg(closes[-60:]) if len(closes) >= 60 else avg(closes)
    vol5 = avg(vols[-5:])

    score = 0
    tags = []

    if open_price > ma5:
        score += 5
        tags.append("open_above_ma5")
    if open_price > ma10:
        score += 5
        tags.append("open_above_ma10")

    short_ma_confirmed = open_price > ma5 and open_price > ma10
    close_breaks_ma5 = close > ma5 and prev.get("close", 0) <= ma5
    close_near_ma10 = close >= ma10 * 0.98 if ma10 else False
    if not short_ma_confirmed:
        if relaxed and close_breaks_ma5 and close_near_ma10:
            score += 4
            tags.append("close_breaks_ma5")
            if close >= ma10:
                score += 2
                tags.append("close_above_ma10")
            else:
                tags.append("close_near_ma10")
        else:
            blockers.append("not_above_short_ma")

    ma20_distance = (ma20 - open_price) / open_price * 100 if open_price else 999
    ma60_distance = (ma60 - open_price) / open_price * 100 if open_price else 999
    if open_price >= ma20:
        score += 5
        tags.append("open_above_ma20")
    elif 0 <= ma20_distance <= 3:
        score += 4
        tags.append("near_ma20_breakout")
    if open_price >= ma60:
        score += 5
        tags.append("open_above_ma60")
    elif 0 <= ma60_distance <= 5:
        score += 4
        tags.append("near_ma60_breakout")

    ma5_distance_pct = (close - ma5) / ma5 * 100 if ma5 else 999
    ma20_above_distance_pct = (close - ma20) / ma20 * 100 if ma20 else 999
    max_ma5_distance = 8.0 if relaxed else settings.pre_breakout_max_ma5_distance_pct
    max_ma20_distance = 15.0 if relaxed else settings.pre_breakout_max_ma20_distance_pct
    if ma5_distance_pct > max_ma5_distance:
        blockers.append("too_far_from_ma5")
    if ma20_above_distance_pct > max_ma20_distance:
        blockers.append("extended_from_ma20")
    if len(history) >= 10 and history[-10].get("close"):
        recent_return = (close - history[-10]["close"]) / history[-10]["close"] * 100
        max_recent_return = 25 if relaxed else 18
        if recent_return > max_recent_return:
            blockers.append("recent_runup_too_high")
    if len(history) >= 20 and history[-20].get("close"):
        return_20d = (close - history[-20]["close"]) / history[-20]["close"] * 100
        max_20d_return = 25.0 if relaxed else settings.pre_breakout_max_20d_return_pct
        if return_20d > max_20d_return:
            blockers.append("twenty_day_runup_too_high")

    prev_high_distance = (prev["high"] - open_price) / open_price * 100 if open_price else 999
    above_prev_high_pct = (open_price - prev["high"]) / prev["high"] * 100 if prev["high"] else 0
    if open_price > prev["high"]:
        max_above_prev_high = 4.0 if relaxed else settings.pre_breakout_max_above_prev_high_pct
        if above_prev_high_pct > max_above_prev_high:
            blockers.append("breakout_too_extended")
        score += 4
        tags.append("just_breaks_prev_high")
    elif 0 <= prev_high_distance <= 3:
        score += 6
        tags.append("near_prev_high_setup")
    elif prev_high_distance > 5:
        blockers.append("too_far_from_prev_high")

    if vol5 and volume > vol5 * 1.5:
        score += 5
        tags.append("volume_expansion")
    elif vol5 and volume > vol5:
        score += 2
        tags.append("volume_above_avg")

    kd_bars = history[-8:] + [
        {
            "date": datetime.now().date().isoformat(),
            "open": row["open"],
            "high": max(row["high"], row["close"]),
            "low": min(row["low"], row["close"]),
            "close": row["close"],
            "volume": row["volume"],
        }
    ]
    k_value, d_value = calc_kd(kd_bars)
    if k_value > 60 and k_value > d_value and d_value > 50:
        score += 8
        tags.append("kd_confirmed")
    elif 30 <= k_value <= 60 and k_value > d_value:
        score += 3
        tags.append("kd_early_turn")
    elif k_value < d_value and close < ma5:
        blockers.append("kd_weak_and_below_ma5")
    kd_gap_limit = 4.0 if relaxed else 3.0
    if k_value > 90 and gap_pct > kd_gap_limit:
        blockers.append("kd_gap_overheat")

    if close >= open_price:
        score += 4
        tags.append("holds_open")
    elif relaxed and close >= open_price * 0.99:
        tags.append("watch_slightly_below_open")
    else:
        blockers.append("falls_below_open")

    return score, tags, blockers


def write_csv(path, rows, fieldnames):
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def localized_value(header, value):
    if value is None:
        return value
    if header in ("modes", "mode"):
        return ",".join(MODE_MAP.get(part.strip(), part.strip()) for part in str(value).split(","))
    if header == "watch_level":
        return WATCH_MAP.get(str(value), value)
    return value


def autosize_sheet(ws, max_width=42):
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        width = 10
        for cell in column[:300]:
            if cell.value is not None:
                width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def style_sheet(ws, header_row=1):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    note_fill = PatternFill("solid", fgColor="FFF2CC")
    note_font = Font(color="7F6000", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(bottom=thin)
    ws.freeze_panes = f"A{header_row + 1}"
    if ws.max_row and ws.max_column:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
    if header_row > 1:
        for cell in ws[1]:
            cell.fill = note_fill
            cell.font = note_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[1].height = 38
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="top")
    autosize_sheet(ws)


def write_xlsx_sheet(wb, title, rows, fields):
    ws = wb.create_sheet(title[:31])
    note = SHEET_NOTES.get(title)
    if note:
        ws.append([f"規則摘要：{note}"] + [None] * (max(1, len(fields)) - 1))
        if len(fields) > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(fields))
    ws.append([HEADER_MAP.get(field, field) for field in fields])
    for row in rows:
        ws.append([localized_value(field, row.get(field)) for field in fields])
    style_sheet(ws, header_row=2 if note else 1)

def write_excel_report(output_dir, snapshot_rows, groups, overlap_rows, excluded, output_name="stock_screen_report.xlsx"):
    wb = Workbook()
    wb.remove(wb.active)

    summary = wb.create_sheet("說明")
    summary.append(["項目", "內容"])
    summary.append(["來源資料夾", str(output_dir)])
    summary.append(["主要分頁", "先看「重疊名單」，再看當沖/隔日沖/週沖"])
    summary.append(["注意", "重疊名單是觀察優先級，不是直接買進訊號"])
    style_sheet(summary)

    overlap_fields = [
        "watch_level",
        "mode_count",
        "code",
        "name",
        "exchange",
        "modes",
        "levels",
        "avg_score",
        "min_score",
        "entry_price",
        "nearest_target_price",
        "nearest_stop_price",
        "min_net_profit",
        "min_rr_ratio",
        "return_pct",
        "volume",
        "reasons",
    ]
    candidate_fields = list(Candidate.__annotations__.keys())
    snapshot_fields = [
        "ts",
        "code",
        "name",
        "exchange",
        "day_trade",
        "category",
        "reference",
        "limit_up",
        "limit_down",
        "open",
        "high",
        "low",
        "close",
        "buy_price",
        "sell_price",
        "volume",
        "total_amount",
    ]
    excluded_fields = list(excluded[0].keys()) if excluded else ["mode", "code", "name", "exclude_reasons"]

    write_xlsx_sheet(wb, "重疊名單", overlap_rows, overlap_fields)
    write_xlsx_sheet(wb, "起漲點A級", [asdict(row) for row in groups["pre_breakout"]], candidate_fields)
    write_xlsx_sheet(wb, "起漲觀察B級", [asdict(row) for row in groups["pre_breakout_watch"]], candidate_fields)
    write_xlsx_sheet(wb, "近期穩健", [asdict(row) for row in groups["stable"]], candidate_fields)
    write_xlsx_sheet(wb, "當沖", [asdict(row) for row in groups["daytrade"]], candidate_fields)
    write_xlsx_sheet(wb, "隔日沖", [asdict(row) for row in groups["overnight"]], candidate_fields)
    write_xlsx_sheet(wb, "週沖", [asdict(row) for row in groups["weekly"]], candidate_fields)
    write_xlsx_sheet(wb, "排除清單", excluded, excluded_fields)
    write_xlsx_sheet(wb, "原始快照", snapshot_rows, snapshot_fields)
    target = output_dir / output_name
    try:
        wb.save(target)
    except PermissionError:
        fallback = output_dir / f"{target.stem}_{datetime.now():%H%M%S}.xlsx"
        wb.save(fallback)
        print(f"xlsx target locked; saved fallback={fallback}")


def labeled_output_names():
    return [
        "pre_breakout_candidates.csv",
        "pre_breakout_candidates.json",
        "pre_breakout_watch_candidates.csv",
        "pre_breakout_watch_candidates.json",
        "overlap_candidates.csv",
        "overlap_candidates.json",
        "daytrade_candidates.csv",
        "daytrade_candidates.json",
        "stable_candidates.csv",
        "stable_candidates.json",
        "overnight_candidates.csv",
        "overnight_candidates.json",
        "weekly_candidates.csv",
        "weekly_candidates.json",
        "candidates.csv",
        "candidates.json",
        "screening_report.md",
    ]


def copy_labeled_outputs(output_dir, label):
    if not label:
        return
    for name in labeled_output_names():
        source = output_dir / name
        if not source.exists():
            continue
        target = source.with_name(f"{source.stem}_{label}{source.suffix}")
        shutil.copy2(source, target)


def cleanup_unlabeled_outputs(output_dir):
    for name in [*labeled_output_names(), "stock_screen_report.xlsx"]:
        path = output_dir / name
        if path.exists():
            path.unlink()


def write_candidates(output_dir, filename, rows):
    dicts = [asdict(row) for row in rows]
    fields = list(Candidate.__annotations__.keys())
    write_csv(output_dir / f"{filename}.csv", dicts, fields)
    (output_dir / f"{filename}.json").write_text(
        json.dumps(dicts, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def build_overlap_candidates(groups):
    by_code = {}
    for mode, rows in groups.items():
        for row in rows:
            item = by_code.setdefault(
                row.code,
                {
                    "code": row.code,
                    "name": row.name,
                    "exchange": row.exchange,
                    "modes": [],
                    "levels": [],
                    "scores": [],
                    "entry_prices": [],
                    "target_prices": [],
                    "stop_prices": [],
                    "net_profits": [],
                    "rr_ratios": [],
                    "return_pct": row.return_pct,
                    "volume": row.volume,
                    "reasons": set(),
                },
            )
            item["modes"].append(mode)
            item["levels"].append(row.level)
            item["scores"].append(row.score)
            item["entry_prices"].append(row.entry_price)
            item["target_prices"].append(row.target_price)
            item["stop_prices"].append(row.stop_price)
            item["net_profits"].append(row.net_profit_at_target)
            item["rr_ratios"].append(row.rr_ratio)
            item["reasons"].update(tag for tag in row.reasons.split(",") if tag)

    overlap = []
    for item in by_code.values():
        mode_count = len(item["modes"])
        if mode_count < 2:
            continue
        watch_level = "core_watch" if mode_count >= 3 else "priority_watch"
        if mode_count >= 4:
            watch_level = "core_watch_with_breakout"
        if mode_count >= 3 and all(level == "A" for level in item["levels"]):
            watch_level = "core_watch_strong"
        overlap.append(
            {
                "watch_level": watch_level,
                "mode_count": mode_count,
                "code": item["code"],
                "name": item["name"],
                "exchange": item["exchange"],
                "modes": ",".join(item["modes"]),
                "levels": ",".join(item["levels"]),
                "avg_score": round(avg(item["scores"]), 1),
                "min_score": round(min(item["scores"]), 1),
                "entry_price": round(avg(item["entry_prices"]), 2),
                "nearest_target_price": round(min(item["target_prices"]), 2),
                "nearest_stop_price": round(max(item["stop_prices"]), 2),
                "min_net_profit": round(min(item["net_profits"]), 0),
                "min_rr_ratio": round(min(item["rr_ratios"]), 2),
                "return_pct": item["return_pct"],
                "volume": item["volume"],
                "reasons": ",".join(sorted(item["reasons"])[:10]),
            }
        )
    overlap.sort(key=lambda row: (row["mode_count"], row["avg_score"], row["min_rr_ratio"]), reverse=True)
    return overlap


def write_overlap(output_dir, rows):
    fields = [
        "watch_level",
        "mode_count",
        "code",
        "name",
        "exchange",
        "modes",
        "levels",
        "avg_score",
        "min_score",
        "entry_price",
        "nearest_target_price",
        "nearest_stop_price",
        "min_net_profit",
        "min_rr_ratio",
        "return_pct",
        "volume",
        "reasons",
    ]
    write_csv(output_dir / "overlap_candidates.csv", rows, fields)
    (output_dir / "overlap_candidates.json").write_text(
        json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def write_report(path, run_time, simulation, settings, groups, overlap_rows, excluded_count):
    lines = [
        f"# Stock Screen Report {run_time:%Y-%m-%d %H:%M:%S}",
        "",
        f"- Simulation: {simulation}",
        f"- Max price: {settings.max_price}",
        f"- Max one-lot amount: {settings.max_lot_amount:.0f}",
        f"- Min risk/reward: {settings.min_rr_ratio}",
        f"- Excluded rows: {excluded_count}",
        "",
        "## Overlap watchlist",
        "",
        "- core_watch: appears in daytrade, overnight, and weekly lists",
        "- priority_watch: appears in two lists",
        "- This is a watchlist, not a buy signal.",
        "",
        "| Rank | Watch | Modes | Code | Name | Avg score | Entry | Near target | Near stop | Min net | Min RR | Return | Volume |",
        "|---:|---|---|---|---|---:|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for index, row in enumerate(overlap_rows[:30], start=1):
        lines.append(
            f"| {index} | {row['watch_level']} | {row['modes']} | {row['code']} | {row['name']} | "
            f"{row['avg_score']:.1f} | {row['entry_price']:.2f} | {row['nearest_target_price']:.2f} | "
            f"{row['nearest_stop_price']:.2f} | {row['min_net_profit']:.0f} | {row['min_rr_ratio']:.2f} | "
            f"{row['return_pct']:.2f}% | {row['volume']} |"
        )
    if not overlap_rows:
        lines.append("| - | No overlap candidates | - | - | - | - | - | - | - | - | - | - | - |")
    lines.append("")

    labels = {
        "pre_breakout": "Pre-breakout A candidates",
        "pre_breakout_watch": "Pre-breakout watch B candidates",
        "stable": "Stable recent candidates",
        "daytrade": "Daytrade candidates",
        "overnight": "Overnight candidates",
        "weekly": "Weekly candidates",
    }
    for mode, rows in groups.items():
        a_rows = [row for row in rows if row.level == "A"]
        lines.extend(
            [
                f"## {labels[mode]}",
                "",
                f"- A candidates: {len(a_rows)}",
                f"- Total candidates: {len(rows)}",
                "",
                "| Rank | Code | Name | Level | Score | Entry | Target | Stop | Net target | RR | Return | Volume | Reasons |",
                "|---:|---|---|---|---:|---:|---:|---:|---:|---:|---:|---:|---|",
            ]
        )
        for index, row in enumerate(rows[:30], start=1):
            lines.append(
                f"| {index} | {row.code} | {row.name} | {row.level} | {row.score:.1f} | "
                f"{row.entry_price:.2f} | {row.target_price:.2f} | {row.stop_price:.2f} | "
                f"{row.net_profit_at_target:.0f} | {row.rr_ratio:.2f} | {row.return_pct:.2f}% | "
                f"{row.volume} | {row.reasons} |"
            )
        if not rows:
            lines.append("| - | - | No candidates | - | - | - | - | - | - | - | - | - | - |")
        lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def rank_snapshot_pool(rows, settings):
    ranked = []
    for row in rows:
        if common_exclusions(row, settings):
            continue
        scores = snapshot_base_scores(row)
        ranked.append((scores["liquidity"] + scores["momentum"] + scores["technical"], row))
    ranked.sort(key=lambda item: item[0], reverse=True)
    return [row for _, row in ranked[: settings.history_pool_size]]


def run_once(output_label=None, skip_non_trading_day=False, after_close=False):
    run_time = datetime.now()
    output_dir = DATA_DIR / run_time.strftime("%Y-%m-%d")
    output_dir.mkdir(parents=True, exist_ok=True)

    if skip_non_trading_day:
        is_trading_day, reason = check_twse_trading_day_after_close(run_time.date())
        if not is_trading_day:
            write_non_trading_day_skip(output_dir, run_time, output_label, reason)
            print(f"skip_non_trading_day date={run_time.date().isoformat()} reason={reason}")
            return
        if reason.startswith("trading_day_check_failed") or reason.startswith("twse_status_unknown"):
            print(f"trading_day_check_warning {reason}", file=sys.stderr)

    api, simulation = load_api()
    settings = load_settings()
    refresh_theme_groups()
    refresh_corporate_action_risks(settings, run_time.date())
    try:
        snapshot_rows, contracts_by_code = fetch_snapshots(api)
        history_pool = rank_snapshot_pool(snapshot_rows, settings)
        history_contracts = [contracts_by_code[row["code"]] for row in history_pool if row["code"] in contracts_by_code]
        history = fetch_history_features(api, history_contracts, settings.kbar_days)
    finally:
        safe_logout(api)

    snapshot_fields = [
        "ts",
        "code",
        "name",
        "exchange",
        "day_trade",
        "category",
        "reference",
        "limit_up",
        "limit_down",
        "open",
        "high",
        "low",
        "close",
        "buy_price",
        "sell_price",
        "volume",
        "total_amount",
    ]
    write_csv(output_dir / "snapshots.csv", snapshot_rows, snapshot_fields)

    theme_features = build_theme_features(snapshot_rows)
    groups = {"pre_breakout": [], "pre_breakout_watch": [], "stable": [], "daytrade": [], "overnight": [], "weekly": []}
    excluded = []
    for row in snapshot_rows:
        for mode in groups:
            candidate, reasons = make_candidate(row, mode, settings, history.get(row["code"]), theme_features, after_close=after_close)
            if candidate:
                groups[mode].append(candidate)
            elif reasons:
                excluded.append({"mode": mode, **row, "exclude_reasons": ",".join(reasons)})

    watch_codes = {row.code for row in groups["pre_breakout_watch"]}
    for candidate in groups["pre_breakout"]:
        if candidate.code not in watch_codes:
            data = asdict(candidate)
            data["mode"] = "pre_breakout_watch"
            data["reasons"] = f"{candidate.reasons},included_from_pre_breakout_a" if candidate.reasons else "included_from_pre_breakout_a"
            groups["pre_breakout_watch"].append(Candidate(**data))
            watch_codes.add(candidate.code)

    for mode in groups:
        groups[mode].sort(key=lambda item: item.score, reverse=True)
        write_candidates(output_dir, f"{mode}_andidates" if False else f"{mode}_candidates", groups[mode])

    overlap_rows = build_overlap_candidates(groups)
    write_overlap(output_dir, overlap_rows)
    write_candidates(output_dir, "candidates", groups["daytrade"])
    write_csv(output_dir / "excluded.csv", excluded, list(excluded[0].keys()) if excluded else ["mode", "code", "name", "exclude_reasons"])
    write_report(output_dir / "screening_report.md", run_time, simulation, settings, groups, overlap_rows, len(excluded))
    if output_label:
        write_excel_report(
            output_dir,
            snapshot_rows,
            groups,
            overlap_rows,
            excluded,
            output_name=f"stock_screen_report_{output_label}.xlsx",
        )
        copy_labeled_outputs(output_dir, output_label)
        cleanup_unlabeled_outputs(output_dir)
    else:
        write_excel_report(output_dir, snapshot_rows, groups, overlap_rows, excluded)


    print(f"output_dir={output_dir}")
    print(
        f"snapshots={len(snapshot_rows)} "
        f"pre_breakout={len(groups['pre_breakout'])} pre_breakout_watch={len(groups['pre_breakout_watch'])} stable={len(groups['stable'])} daytrade={len(groups['daytrade'])} overnight={len(groups['overnight'])} "
        f"weekly={len(groups['weekly'])} overlap={len(overlap_rows)} excluded={len(excluded)}"
    )
    print("[overlap]")
    for row in overlap_rows[:5]:
        print(
            f"{row['watch_level']} {row['code']} {row['name']} modes={row['modes']} "
            f"avg_score={row['avg_score']:.1f} min_rr={row['min_rr_ratio']:.2f}"
        )
    for mode in ("pre_breakout", "pre_breakout_watch", "stable", "daytrade", "overnight", "weekly"):
        print(f"[{mode}]")
        for row in groups[mode][:5]:
            print(
                f"{row.level} {row.code} {row.name} score={row.score:.1f} "
                f"entry={row.entry_price:.2f} target={row.target_price:.2f} rr={row.rr_ratio:.2f}"
            )


def run_window():
    today = datetime.now().date()
    start = datetime.combine(today, dtime(8, 30))
    end = datetime.combine(today, dtime(9, 10))
    now = datetime.now()
    if now < start:
        time.sleep((start - now).total_seconds())
    while datetime.now() < end:
        interval = 30 if datetime.now().time() < dtime(9, 0) else 5
        run_once()
        time.sleep(interval)
    run_once(output_label="0910")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--once", action="store_true", help="Run one current-time screen.")
    parser.add_argument("--window", action="store_true", help="Run 8:30-9:10 window collection.")
    parser.add_argument("--output-label", default=None, help="Also write labeled output copies, for example 0910, 1000, or 1400.")
    parser.add_argument("--skip-non-trading-day", action="store_true", help="Skip after-close runs on weekends or TWSE non-trading days.")
    parser.add_argument("--after-close", action="store_true", help="Use after-close screening rules, including not rejecting missing live bid/ask spread.")
    args = parser.parse_args()
    if args.window:
        run_window()
    else:
        run_once(output_label=args.output_label, skip_non_trading_day=args.skip_non_trading_day, after_close=args.after_close)


if __name__ == "__main__":
    main()







