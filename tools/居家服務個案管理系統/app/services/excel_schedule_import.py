import os
import re
from datetime import date, time as dtime

import openpyxl
from sqlalchemy.orm import Session
from app.models.caregiver_service_record import CaregiverServiceRecord
from app.models.case import Case
from app.models.user import User, UserRole

# 每日欄位配置：(date_col, minutes_col, entry_start_col)
DAY_COLUMNS = {
    6: ("B", "C", "B"),
    0: ("E", "F", "E"),
    1: ("H", "I", "H"),
    2: ("K", "M", "K"),
    3: ("O", "P", "O"),
    4: ("R", "S", "R"),
    5: ("U", "V", "U"),
}

TIME_RANGE_RE = re.compile(r"(\d{1,2}):(\d{2})~(\d{1,2}):(\d{2})")
SERVICE_CODE_RE = re.compile(r"([A-Z]{1,3}\d{1,2}(?:-\d+)?(?:[a-z]\d?)?)\s*x\s*(\d+)", re.IGNORECASE)
DATE_CELL_RE = re.compile(r"(\d{1,2})/(\d{1,2})")
CAREGIVER_RE = re.compile(r"員工姓名[：:]\s*(?:.*?\s*-\s*)?(.+)$")


def parse_caregiver_name(row2: str | None) -> str:
    if not row2:
        return ""
    m = CAREGIVER_RE.search(row2)
    return m.group(1).strip() if m else row2.strip()


def _is_date_cell(val) -> bool:
    if isinstance(val, str) and DATE_CELL_RE.match(val.strip()):
        return True
    if isinstance(val, date):
        return True
    return False


def _to_date(val, year: int = 2026) -> date | None:
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        m = DATE_CELL_RE.search(val.strip())
        if m:
            return date(year, int(m.group(1)), int(m.group(2)))
    return None


def _parse_entry_text(text: str, service_date: date) -> list[dict]:
    if not text or not text.strip() or text.strip() in (" ", ""):
        return []

    raw = text.strip()
    notes = re.findall(r"\(([^)]*)\)", raw)
    clean = re.sub(r"\([^)]*\)", "", raw).strip()
    lines = [l.strip() for l in clean.split("\n") if l.strip()]

    results = []
    for line in lines:
        tm = TIME_RANGE_RE.search(line)
        if not tm:
            continue
        start_h, start_m, end_h, end_m = map(int, tm.groups())
        st = dtime(start_h, start_m)
        et = dtime(end_h, end_m)
        total_min = (end_h * 60 + end_m) - (start_h * 60 + start_m)
        if total_min <= 0:
            continue

        after_time = line[tm.end() :].strip()
        cm = SERVICE_CODE_RE.search(after_time)
        if cm:
            case_name = after_time[: cm.start()].strip()
            codes_text = after_time[cm.start() :]
        else:
            case_name = after_time.strip()
            codes_text = ""

        codes = []
        for sm in SERVICE_CODE_RE.finditer(codes_text):
            codes.append(f"{sm.group(1)}x{sm.group(2)}")

        is_leave = any(
            kw in n for n in notes for kw in ("假", "停", "取消", "暫停")
        )

        results.append(
            {
                "start_time": st,
                "end_time": et,
                "minutes": total_min,
                "case_name_raw": case_name,
                "service_codes": ", ".join(codes) if codes else "",
                "note": "; ".join(notes) if notes else None,
                "is_leave": is_leave,
            }
        )

    return results


def parse_file(filepath: str, year: int = 2026) -> list[dict]:
    """Parse one Rpt_EmpScheduleCalendarMultiplex file into schedule entries."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    caregiver_name = parse_caregiver_name(ws["A2"].value)
    entries = []

    # Iterate to find date rows and their detail rows
    for row_num in range(6, ws.max_row + 1):
        for weekday, (dcol, _, ecol) in DAY_COLUMNS.items():
            cell = ws[dcol + str(row_num)]
            if not _is_date_cell(cell.value):
                continue
            svc_date = _to_date(cell.value, year)
            if not svc_date:
                continue

            day_entries = []
            # Collect entries from subsequent rows for this day column
            for detail_row in range(row_num + 1, ws.max_row + 1):
                ecell = ws[ecol + str(detail_row)]
                # If we hit another date row, stop
                if _is_date_cell(ecell.value) and ecol == dcol:
                    break
                # Check if this row has a detail in this day column
                detail_cell = ws[ecol + str(detail_row)]
                if detail_cell.value is not None:
                    parsed = _parse_entry_text(str(detail_cell.value), svc_date)
                    day_entries.extend(parsed)

            for e in day_entries:
                e["caregiver_name_raw"] = caregiver_name
                e["service_date"] = svc_date
                entries.append(e)

    wb.close()
    return entries


def _match_case_name(name: str, db: Session):
    """Fuzzy match a case name from Excel to a case in DB."""
    name = name.strip()
    case = db.query(Case).filter(Case.name == name).first()
    if case:
        return case.id, name
    case = db.query(Case).filter(Case.name.like(f"%{name}%")).first()
    if case:
        return case.id, case.name
    case = db.query(Case).filter(Case.name.like(f"%{name[:2]}%")).first()
    if case:
        return case.id, case.name
    return None, name


def _match_caregiver_name(name: str, db: Session):
    name = name.strip()
    u = db.query(User).filter(User.role == UserRole.caregiver, User.display_name == name).first()
    if u:
        return u.id, name
    u = db.query(User).filter(User.role == UserRole.caregiver, User.display_name.like(f"%{name}%")).first()
    if u:
        return u.id, u.display_name
    return None, name


def import_entries_to_db(entries: list[dict], month: date, db: Session) -> dict:
    """Import parsed schedule entries into CaregiverServiceRecord, replacing month data."""
    from datetime import timedelta
    month_start = month
    if month.month == 12:
        month_end = date(month.year, 12, 31)
    else:
        month_end = date(month.year, month.month + 1, 1) - timedelta(days=1)

    db.query(CaregiverServiceRecord).filter(
        CaregiverServiceRecord.service_date >= month_start,
        CaregiverServiceRecord.service_date <= month_end,
    ).delete(synchronize_session=False)

    added = 0
    skipped_no_case = 0
    skipped_no_cg = 0

    for e in entries:
        cg_id, cg_name = _match_caregiver_name(e["caregiver_name_raw"], db)
        if not cg_id:
            skipped_no_cg += 1
            continue
        case_id, case_name = _match_case_name(e["case_name_raw"], db)
        if not case_id:
            skipped_no_case += 1
            continue

        status = "leave" if e.get("is_leave") else "external_import"

        db.add(CaregiverServiceRecord(
            case_id=case_id,
            caregiver_id=cg_id,
            service_date=e["service_date"],
            start_time=e["start_time"],
            end_time=e["end_time"],
            minutes=e["minutes"],
            case_name_raw=case_name,
            caregiver_name_raw=cg_name,
            service_codes=e.get("service_codes", ""),
            formalization_status=status,
            funding_source="補助",
            note=e.get("note"),
        ))
        added += 1

    db.commit()
    return {"added": added, "skipped_no_case": skipped_no_case, "skipped_no_cg": skipped_no_cg}


def parse_directory(dirpath: str, year: int = 2026) -> dict[str, list[dict]]:
    """Parse all xlsx files in a directory, keyed by caregiver name."""
    result = {}
    for fname in sorted(os.listdir(dirpath)):
        if not fname.endswith(".xlsx") or not fname.startswith("Rpt_EmpSchedule"):
            continue
        fpath = os.path.join(dirpath, fname)
        try:
            entries = parse_file(fpath, year)
            if entries:
                cg_name = entries[0]["caregiver_name_raw"]
                result[cg_name] = entries
        except Exception as e:
            print(f"Error parsing {fname}: {e}")
    return result
